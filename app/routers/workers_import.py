from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import logging
import re
from typing import Any, Optional

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session

from .. import models, schemas
from ..config.config import get_db
from ..security import WRITE_RH_ROLES, can_manage_worker, require_roles
from ..services.audit_service import record_audit
from ..services.data_mapping_service import DataMappingError, map_user_excel_to_template
from ..services.master_data_service import sync_worker_master_data
from ..services.tabular_io import dataframe_to_csv_bytes, issues_to_csv, normalize_header, read_tabular_bytes

router = APIRouter(prefix="/workers/import", tags=["workers-import"])
logger = logging.getLogger(__name__)

WORKER_TEMPLATE_COLUMNS = [
    "Raison Sociale",
    "Matricule",
    "Nom",
    "Prenom",
    "Sexe (M/F)",
    "Date de Naissance (JJ/MM/AAAA)",
    "Lieu de Naissance",
    "Situation Familiale",
    "Nombre Enfants",
    "CIN",
    "CIN Delivre le (JJ/MM/AAAA)",
    "CIN Lieu de delivrance",
    "Adresse",
    "Telephone",
    "Email",
    "Numero CNaPS",
    "Date Embauche (JJ/MM/AAAA)",
    "Date Debut Contrat (JJ/MM/AAAA)",
    "Date Fin Contrat (JJ/MM/AAAA)",
    "Nature du Contrat",
    "Duree Essai (jours)",
    "Date Fin Essai (JJ/MM/AAAA)",
    "Poste Actuel",
    "Categorie Professionnelle",
    "Indice Classification",
    "Etablissement",
    "Departement",
    "Service",
    "Unite",
    "Type Regime (Agricole/Non Agricole)",
    "Secteur (agricole/non_agricole)",
    "Horaire Hebdo",
    "Salaire Base",
    "Taux Horaire",
    "Mode de Paiement",
    "RIB",
    "Nom de la Banque",
    "Nom du Guichet",
    "BIC / SWIFT",
    "Code Banque",
    "Code Guichet",
    "Numero de Compte",
    "Cle RIB",
    "SMIE Agence",
    "SMIE Numero Carte",
    "Avantage Vehicule",
    "Avantage Logement",
    "Avantage Telephone",
    "Avantage Autres",
    "Solde Conge Initial",
    "Niveau Etudes Principal",
    "Autres Diplomes",
    "Annees Experience",
    "Competences (separees par ;)",
    "Langues (separees par ;)",
    "Source Recrutement",
    "Type Recrutement",
    "Statut Onboarding",
    "Reference Candidat",
    "Observations",
]

REQUIRED_COLUMNS = ["Matricule", "Nom"]
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

COLUMN_ALIASES = {
    "Prenom": ["Prénom", "Prenoms", "First Name"],
    "Matricule": ["Matricules", "Employee ID", "ID salarie"],
    "Nom": ["Noms", "Last Name", "Nom de famille"],
    "Date Embauche (JJ/MM/AAAA)": ["Date Embauche", "Date d'embauche", "Date entree"],
    "Date Debut Contrat (JJ/MM/AAAA)": ["Date Debut Contrat", "Date debut"],
    "Date Fin Contrat (JJ/MM/AAAA)": ["Date Fin Contrat", "Date fin"],
    "Nature du Contrat": ["Type Contrat", "Contract Type"],
    "Categorie Professionnelle": ["Categorie", "Classification", "CSP"],
    "Indice Classification": ["Indice", "Classification Index"],
    "Salaire Base": ["Salaire de base", "Base Salary"],
    "Taux Horaire": ["Salaire Horaire", "Hourly Rate"],
    "Type Regime (Agricole/Non Agricole)": ["Type Regime", "Regime"],
    "Numero CNaPS": ["Numero CNAPS", "CNaPS"],
    "Raison Sociale": ["Employeur", "Employer"],
    "Competences (separees par ;)": ["Competences", "Skills"],
    "Langues (separees par ;)": ["Langues", "Languages"],
    "Source Recrutement": ["Source"],
    "Type Recrutement": ["Recruitment Type"],
    "Statut Onboarding": ["Onboarding Status"],
    "Reference Candidat": ["Candidate ID", "Candidate Reference"],
}

EXAMPLE_ROW = {
    "Raison Sociale": "Karibo Services",
    "Matricule": "M001",
    "Nom": "RAKOTO",
    "Prenom": "Jean",
    "Sexe (M/F)": "M",
    "Date de Naissance (JJ/MM/AAAA)": "15/05/1985",
    "Lieu de Naissance": "Antananarivo",
    "Situation Familiale": "Marie(e)",
    "Nombre Enfants": 2,
    "CIN": "101234567890",
    "CIN Delivre le (JJ/MM/AAAA)": "01/01/2020",
    "CIN Lieu de delivrance": "Antananarivo",
    "Adresse": "Lot IVC Tana",
    "Telephone": "0340000000",
    "Email": "jean.rakoto@example.com",
    "Numero CNaPS": "9876543210X",
    "Date Embauche (JJ/MM/AAAA)": "01/01/2024",
    "Date Debut Contrat (JJ/MM/AAAA)": "01/01/2024",
    "Date Fin Contrat (JJ/MM/AAAA)": "",
    "Nature du Contrat": "CDI",
    "Duree Essai (jours)": 90,
    "Date Fin Essai (JJ/MM/AAAA)": "31/03/2024",
    "Poste Actuel": "Gestionnaire Paie",
    "Categorie Professionnelle": "M1",
    "Indice Classification": "1A",
    "Etablissement": "Siege Social",
    "Departement": "Ressources Humaines",
    "Service": "Administration",
    "Unite": "Paie",
    "Type Regime (Agricole/Non Agricole)": "Non Agricole",
    "Secteur (agricole/non_agricole)": "non_agricole",
    "Horaire Hebdo": 40,
    "Salaire Base": 250000,
    "Taux Horaire": 1442.31,
    "Mode de Paiement": "Virement",
    "RIB": "",
    "Nom de la Banque": "BNI",
    "Nom du Guichet": "Antananarivo Centre",
    "BIC / SWIFT": "BNIMMG",
    "Code Banque": "00005",
    "Code Guichet": "00081",
    "Numero de Compte": "12345678901",
    "Cle RIB": "63",
    "SMIE Agence": "Tana",
    "SMIE Numero Carte": "",
    "Avantage Vehicule": 0,
    "Avantage Logement": 0,
    "Avantage Telephone": 0,
    "Avantage Autres": 0,
    "Solde Conge Initial": 0,
    "Niveau Etudes Principal": "Licence",
    "Autres Diplomes": "Certification paie",
    "Annees Experience": 5,
    "Competences (separees par ;)": "Paie;Excel avance",
    "Langues (separees par ;)": "Francais;Malagasy",
    "Source Recrutement": "Cooptation",
    "Type Recrutement": "Interne",
    "Statut Onboarding": "hired",
    "Reference Candidat": "",
    "Observations": "Dossier complet",
}


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.strip() == "":
        return False
    return not pd.isna(value)


def _safe_str(value: Any) -> str:
    return str(value).strip() if _has_value(value) else ""


def _safe_float(value: Any, default: float = 0.0) -> float:
    if not _has_value(value):
        return default
    try:
        return float(value)
    except Exception:
        return default


def _safe_int(value: Any, default: int = 0) -> int:
    if not _has_value(value):
        return default
    try:
        return int(float(value))
    except Exception:
        return default


def _parse_optional_date(value: Any) -> Optional[date]:
    if not _has_value(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    try:
        return pd.to_datetime(raw).date()
    except Exception:
        return None


def _parse_optional_date_strict(value: Any) -> tuple[Optional[date], bool]:
    if not _has_value(value):
        return None, True
    parsed = _parse_optional_date(value)
    return parsed, parsed is not None


def _parse_optional_float_strict(value: Any, default: float = 0.0) -> tuple[float, bool]:
    if not _has_value(value):
        return default, True
    try:
        parsed = float(str(value).strip().replace(" ", "").replace(",", "."))
        return parsed, True
    except Exception:
        return default, False


def _parse_optional_int_strict(value: Any, default: int = 0) -> tuple[int, bool]:
    if not _has_value(value):
        return default, True
    try:
        parsed = int(float(str(value).strip().replace(" ", "").replace(",", ".")))
        return parsed, True
    except Exception:
        return default, False


def _normalize_phone(value: Any) -> tuple[str, bool]:
    raw = _safe_str(value)
    if not raw:
        return "", True
    normalized = raw.replace(" ", "").replace("-", "").replace(".", "")
    if normalized.startswith("00"):
        normalized = f"+{normalized[2:]}"
    allowed = set("+()0123456789")
    if any(char not in allowed for char in normalized):
        return raw, False
    digits_count = sum(char.isdigit() for char in normalized)
    if digits_count < 7:
        return raw, False
    return normalized, True


def _normalize_email(value: Any) -> tuple[str, bool]:
    raw = _safe_str(value).lower()
    if not raw:
        return "", True
    return raw, bool(EMAIL_RE.match(raw))


def _normalize_cin(value: Any) -> str:
    raw = _safe_str(value).upper()
    if not raw:
        return ""
    return "".join(char for char in raw if char.isalnum())


def _split_multi_values(value: Any) -> list[str]:
    raw = _safe_str(value).replace("|", ";").replace(",", ";").replace("\n", ";")
    return [item.strip() for item in raw.split(";") if item.strip()]


def _normalize_skill_code(value: str) -> str:
    token = "".join(char if char.isalnum() else "_" for char in value.lower())
    token = "_".join(part for part in token.split("_") if part)
    return (token or "competence")[:80]


def _build_column_mapping_with_aliases(actual_columns: list[object]) -> tuple[dict[str, str], list[str], list[str]]:
    alias_lookup: dict[str, str] = {}
    for canonical in WORKER_TEMPLATE_COLUMNS:
        alias_lookup[normalize_header(canonical)] = canonical
        for alias in COLUMN_ALIASES.get(canonical, []):
            alias_lookup[normalize_header(alias)] = canonical

    mapping: dict[str, str] = {}
    unknown: list[str] = []
    for raw in actual_columns:
        raw_str = str(raw).strip()
        canonical = alias_lookup.get(normalize_header(raw_str))
        if canonical:
            mapping.setdefault(canonical, raw_str)
        else:
            unknown.append(raw_str)
    missing = [required for required in REQUIRED_COLUMNS if required not in mapping]
    return mapping, unknown, missing


def _build_workers_template_xlsx(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Salaries"
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(WORKER_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    ws.freeze_panes = "A2"
    for idx, column in enumerate(WORKER_TEMPLATE_COLUMNS, start=1):
        width = min(max(len(column) + 2, 12), 50)
        ws.column_dimensions[get_column_letter(idx)].width = width

    instructions = wb.create_sheet("Instructions")
    lines = [
        ("Regle", "Matricule et Nom sont obligatoires."),
        ("Regle", "Les champs non essentiels peuvent rester vides."),
        ("Regle", "Les colonnes inconnues sont ignorees."),
        ("Flux", "Import -> Worker -> MasterData (auto sync)."),
        ("Flux", "Recrutement/Talents sont enrichis si colonnes remplies."),
    ]
    for row_idx, (col1, col2) in enumerate(lines, start=1):
        instructions.cell(row=row_idx, column=1, value=col1)
        instructions.cell(row=row_idx, column=2, value=col2)
    instructions.column_dimensions["A"].width = 16
    instructions.column_dimensions["B"].width = 96

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _add_issue(
    issues: list[dict[str, Any]],
    *,
    row_number: int,
    code: str,
    message: str,
    column: Optional[str] = None,
    value: Optional[Any] = None,
) -> None:
    issues.append(
        {
            "row_number": row_number,
            "column": column,
            "code": code,
            "message": message,
            "value": None if value is None else str(value),
        }
    )


def _build_report(
    *,
    mode: str,
    total_rows: int,
    processed_rows: int,
    created: int,
    updated: int,
    skipped: int,
    failed: int,
    unknown_columns: list[str],
    missing_columns: list[str],
    issues: list[dict[str, Any]],
) -> schemas.TabularImportReport:
    return schemas.TabularImportReport(
        mode=mode if mode in {"create", "update", "mixed"} else "mixed",
        total_rows=total_rows,
        processed_rows=processed_rows,
        created=created,
        updated=updated,
        skipped=skipped,
        failed=failed,
        unknown_columns=unknown_columns,
        missing_columns=missing_columns,
        issues=[schemas.ImportIssue(**item) for item in issues],
        error_report_csv=issues_to_csv(issues) if issues else None,
    )


def _ensure_talent_skill(db: Session, employer_id: int, name: str) -> models.TalentSkill:
    existing = (
        db.query(models.TalentSkill)
        .filter(models.TalentSkill.employer_id == employer_id, func.lower(models.TalentSkill.name) == name.lower())
        .first()
    )
    if existing:
        return existing
    code_base = _normalize_skill_code(name)
    code = code_base
    n = 1
    while (
        db.query(models.TalentSkill)
        .filter(models.TalentSkill.employer_id == employer_id, models.TalentSkill.code == code)
        .first()
        is not None
    ):
        n += 1
        suffix = f"_{n}"
        code = f"{code_base[: max(1, 80 - len(suffix))]}{suffix}"
    skill = models.TalentSkill(
        employer_id=employer_id,
        code=code,
        name=name,
        description="Cree via import travailleurs",
        scale_max=5,
        is_active=True,
    )
    db.add(skill)
    db.flush()
    return skill


def _upsert_worker_skills(db: Session, worker: models.Worker, names: list[str]) -> None:
    for skill_name in sorted(set(name for name in names if name)):
        skill = _ensure_talent_skill(db, worker.employer_id, skill_name)
        existing_link = (
            db.query(models.TalentEmployeeSkill)
            .filter(models.TalentEmployeeSkill.worker_id == worker.id, models.TalentEmployeeSkill.skill_id == skill.id)
            .first()
        )
        if existing_link:
            existing_link.source = "import_excel"
            if existing_link.level <= 0:
                existing_link.level = 3
            continue
        db.add(models.TalentEmployeeSkill(worker_id=worker.id, skill_id=skill.id, level=3, source="import_excel"))


def _upsert_recruitment_candidate(
    *,
    db: Session,
    employer_id: int,
    first_name: str,
    last_name: str,
    email: str,
    phone: str,
    education_level: str,
    experience_years: float,
    source_recruitment: str,
    recruitment_type: str,
    onboarding_status: str,
    candidate_reference: str,
    summary: str,
) -> Optional[models.RecruitmentCandidate]:
    has_payload = any(
        [
            email,
            education_level,
            source_recruitment,
            recruitment_type,
            onboarding_status,
            candidate_reference,
            summary,
        ]
    )
    if not has_payload:
        return None

    candidate: Optional[models.RecruitmentCandidate] = None
    if candidate_reference.isdigit():
        candidate = (
            db.query(models.RecruitmentCandidate)
            .filter(
                models.RecruitmentCandidate.id == int(candidate_reference),
                models.RecruitmentCandidate.employer_id == employer_id,
            )
            .first()
        )
    if candidate is None and email:
        candidate = (
            db.query(models.RecruitmentCandidate)
            .filter(
                models.RecruitmentCandidate.employer_id == employer_id,
                func.lower(models.RecruitmentCandidate.email) == email.lower(),
            )
            .first()
        )

    merged_source = " / ".join(part for part in [source_recruitment, recruitment_type] if part)
    status = onboarding_status or "hired"
    if candidate is None:
        if not email:
            return None
        candidate = models.RecruitmentCandidate(
            employer_id=employer_id,
            first_name=first_name or "N/A",
            last_name=last_name or "N/A",
            email=email,
            phone=phone or None,
            education_level=education_level or None,
            experience_years=experience_years,
            source=merged_source or None,
            status=status,
            summary=summary or None,
        )
        db.add(candidate)
        db.flush()
        return candidate

    if first_name:
        candidate.first_name = first_name
    if last_name:
        candidate.last_name = last_name
    if email:
        candidate.email = email
    if phone:
        candidate.phone = phone
    if education_level:
        candidate.education_level = education_level
    if experience_years > 0:
        candidate.experience_years = experience_years
    if merged_source:
        candidate.source = merged_source
    if status:
        candidate.status = status
    if summary:
        candidate.summary = summary
    return candidate


def _build_template_dataframe(
    db: Session,
    user: models.AppUser,
    *,
    prefilled: bool,
    employer_id: Optional[int],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if not prefilled:
        rows.append(dict(EXAMPLE_ROW))
        return pd.DataFrame(rows, columns=WORKER_TEMPLATE_COLUMNS)

    query = db.query(models.Worker).order_by(models.Worker.matricule.asc())
    if employer_id is not None:
        if not can_manage_worker(db, user, employer_id=employer_id):
            raise HTTPException(status_code=403, detail="Forbidden")
        query = query.filter(models.Worker.employer_id == employer_id)
    workers = query.all()

    employer_ids = {worker.employer_id for worker in workers}
    employers = db.query(models.Employer).filter(models.Employer.id.in_(employer_ids)).all() if employer_ids else []
    employer_map = {item.id: item for item in employers}

    worker_ids = [worker.id for worker in workers]
    worker_skills: dict[int, list[str]] = {}
    if worker_ids:
        links = db.query(models.TalentEmployeeSkill).filter(models.TalentEmployeeSkill.worker_id.in_(worker_ids)).all()
        skill_ids = {link.skill_id for link in links}
        skills = db.query(models.TalentSkill).filter(models.TalentSkill.id.in_(skill_ids)).all() if skill_ids else []
        skill_map = {skill.id: skill.name for skill in skills}
        for link in links:
            skill_name = skill_map.get(link.skill_id)
            if skill_name:
                worker_skills.setdefault(link.worker_id, []).append(skill_name)

    candidate_map: dict[tuple[int, str], models.RecruitmentCandidate] = {}
    email_keys = {(worker.employer_id, (worker.email or "").lower().strip()) for worker in workers if worker.email}
    if email_keys:
        candidates = (
            db.query(models.RecruitmentCandidate)
            .filter(
                models.RecruitmentCandidate.employer_id.in_(list({item[0] for item in email_keys})),
                func.lower(models.RecruitmentCandidate.email).in_(list({item[1] for item in email_keys})),
            )
            .all()
        )
        for candidate in candidates:
            candidate_map[(candidate.employer_id, (candidate.email or "").lower().strip())] = candidate

    for worker in workers:
        if not can_manage_worker(db, user, worker=worker):
            continue
        employer = employer_map.get(worker.employer_id)
        regime = db.query(models.TypeRegime).filter(models.TypeRegime.id == worker.type_regime_id).first()
        regime_label = "Agricole" if regime and regime.code == "agricole" else "Non Agricole"
        candidate = candidate_map.get((worker.employer_id, (worker.email or "").lower().strip()), None)

        rows.append(
            {
                "Raison Sociale": employer.raison_sociale if employer else "",
                "Matricule": worker.matricule or "",
                "Nom": worker.nom or "",
                "Prenom": worker.prenom or "",
                "Sexe (M/F)": worker.sexe or "",
                "Date de Naissance (JJ/MM/AAAA)": worker.date_naissance.strftime("%d/%m/%Y") if worker.date_naissance else "",
                "Lieu de Naissance": worker.lieu_naissance or "",
                "Situation Familiale": worker.situation_familiale or "",
                "Nombre Enfants": worker.nombre_enfant or 0,
                "CIN": worker.cin or "",
                "CIN Delivre le (JJ/MM/AAAA)": worker.cin_delivre_le.strftime("%d/%m/%Y") if worker.cin_delivre_le else "",
                "CIN Lieu de delivrance": worker.cin_lieu or "",
                "Adresse": worker.adresse or "",
                "Telephone": worker.telephone or "",
                "Email": worker.email or "",
                "Numero CNaPS": worker.cnaps_num or "",
                "Date Embauche (JJ/MM/AAAA)": worker.date_embauche.strftime("%d/%m/%Y") if worker.date_embauche else "",
                "Date Debut Contrat (JJ/MM/AAAA)": worker.date_embauche.strftime("%d/%m/%Y") if worker.date_embauche else "",
                "Date Fin Contrat (JJ/MM/AAAA)": worker.date_debauche.strftime("%d/%m/%Y") if worker.date_debauche else "",
                "Nature du Contrat": worker.nature_contrat or "",
                "Duree Essai (jours)": worker.duree_essai_jours or 0,
                "Date Fin Essai (JJ/MM/AAAA)": worker.date_fin_essai.strftime("%d/%m/%Y") if worker.date_fin_essai else "",
                "Poste Actuel": worker.poste or "",
                "Categorie Professionnelle": worker.categorie_prof or "",
                "Indice Classification": worker.indice or "",
                "Etablissement": worker.etablissement or "",
                "Departement": worker.departement or "",
                "Service": worker.service or "",
                "Unite": worker.unite or "",
                "Type Regime (Agricole/Non Agricole)": regime_label,
                "Secteur (agricole/non_agricole)": worker.secteur or "",
                "Horaire Hebdo": worker.horaire_hebdo or 40,
                "Salaire Base": worker.salaire_base or 0,
                "Taux Horaire": worker.salaire_horaire or 0,
                "Mode de Paiement": worker.mode_paiement or "",
                "RIB": worker.rib or "",
                "Nom de la Banque": worker.banque or "",
                "Nom du Guichet": worker.nom_guichet or "",
                "BIC / SWIFT": worker.bic or "",
                "Code Banque": worker.code_banque or "",
                "Code Guichet": worker.code_guichet or "",
                "Numero de Compte": worker.compte_num or "",
                "Cle RIB": worker.cle_rib or "",
                "SMIE Agence": worker.smie_agence or "",
                "SMIE Numero Carte": worker.smie_carte_num or "",
                "Avantage Vehicule": worker.avantage_vehicule or 0,
                "Avantage Logement": worker.avantage_logement or 0,
                "Avantage Telephone": worker.avantage_telephone or 0,
                "Avantage Autres": worker.avantage_autres or 0,
                "Solde Conge Initial": worker.solde_conge_initial or 0,
                "Niveau Etudes Principal": candidate.education_level if candidate else "",
                "Autres Diplomes": "",
                "Annees Experience": candidate.experience_years if candidate else 0,
                "Competences (separees par ;)": "; ".join(sorted(set(worker_skills.get(worker.id, [])))),
                "Langues (separees par ;)": "",
                "Source Recrutement": candidate.source if candidate else "",
                "Type Recrutement": "",
                "Statut Onboarding": candidate.status if candidate else "",
                "Reference Candidat": candidate.id if candidate else "",
                "Observations": candidate.summary if candidate else "",
            }
        )
    return pd.DataFrame(rows, columns=WORKER_TEMPLATE_COLUMNS)


def _import_workers_dataframe(
    *,
    df: pd.DataFrame,
    update_existing: bool,
    db: Session,
    user: models.AppUser,
    dry_run: bool,
) -> tuple[schemas.TabularImportReport, int, int, int]:
    mode = "mixed" if update_existing else "create"
    mapping, unknown_columns, missing_columns = _build_column_mapping_with_aliases(df.columns.tolist())
    issues: list[dict[str, Any]] = []

    if unknown_columns:
        _add_issue(
            issues,
            row_number=1,
            code="unknown_columns_ignored",
            message=f"Colonnes inconnues ignorees: {', '.join(unknown_columns)}",
        )
    if missing_columns:
        _add_issue(
            issues,
            row_number=1,
            code="missing_columns",
            message=f"Colonnes obligatoires manquantes: {', '.join(missing_columns)}",
        )
        report = _build_report(
            mode=mode,
            total_rows=0,
            processed_rows=0,
            created=0,
            updated=0,
            skipped=0,
            failed=0,
            unknown_columns=unknown_columns,
            missing_columns=missing_columns,
            issues=issues,
        )
        return report, 0, 0, 0

    regimes = db.query(models.TypeRegime).all()
    default_regime = next((item for item in regimes if item.code == "non_agricole"), regimes[0] if regimes else None)
    if not default_regime:
        raise HTTPException(status_code=400, detail="Aucun regime configure.")
    employers = db.query(models.Employer).all()
    if not employers:
        raise HTTPException(status_code=400, detail="Aucun employeur configure.")
    employer_by_name = {item.raison_sociale.lower().strip(): item for item in employers}
    default_employer = employers[0]

    total_rows = 0
    processed_rows = 0
    created = 0
    updated = 0
    skipped = 0
    failed = 0
    seen_matricules: set[str] = set()
    seen_cins: set[str] = set()
    seen_emails: set[str] = set()

    def value(row: pd.Series, column: str) -> Any:
        source_column = mapping.get(column)
        return row.get(source_column) if source_column else None

    for idx, row in df.iterrows():
        row_number = idx + 2
        raw_matricule = value(row, "Matricule")
        raw_nom = value(row, "Nom")
        if not _has_value(raw_matricule) and not _has_value(raw_nom):
            continue
        if str(raw_matricule).strip() == "M001" and str(raw_nom).strip().upper() == "RAKOTO":
            skipped += 1
            _add_issue(issues, row_number=row_number, code="template_example_skipped", message="Ligne exemple ignoree.")
            continue

        total_rows += 1
        matricule = _safe_str(raw_matricule)
        if not matricule:
            failed += 1
            _add_issue(issues, row_number=row_number, column="Matricule", code="missing_matricule", message="Matricule obligatoire.")
            continue
        if matricule in seen_matricules:
            failed += 1
            _add_issue(
                issues,
                row_number=row_number,
                column="Matricule",
                code="duplicate_in_file",
                message=f"Matricule duplique dans le fichier: {matricule}.",
                value=matricule,
            )
            continue
        seen_matricules.add(matricule)

        nom = _safe_str(raw_nom).upper()
        if not nom:
            failed += 1
            _add_issue(issues, row_number=row_number, column="Nom", code="missing_name", message=f"Nom obligatoire pour {matricule}.")
            continue

        email, email_valid = _normalize_email(value(row, "Email"))
        if not email_valid:
            failed += 1
            _add_issue(
                issues,
                row_number=row_number,
                column="Email",
                code="invalid_email",
                message=f"Email invalide pour {matricule}.",
                value=value(row, "Email"),
            )
            continue
        if email:
            if email in seen_emails:
                failed += 1
                _add_issue(
                    issues,
                    row_number=row_number,
                    column="Email",
                    code="duplicate_email_in_file",
                    message=f"Email duplique dans le fichier: {email}.",
                    value=email,
                )
                continue
            seen_emails.add(email)

        cin = _normalize_cin(value(row, "CIN"))
        if cin:
            if len(cin) < 6:
                failed += 1
                _add_issue(
                    issues,
                    row_number=row_number,
                    column="CIN",
                    code="invalid_cin",
                    message=f"CIN invalide pour {matricule}.",
                    value=cin,
                )
                continue
            if cin in seen_cins:
                failed += 1
                _add_issue(
                    issues,
                    row_number=row_number,
                    column="CIN",
                    code="duplicate_cin_in_file",
                    message=f"CIN duplique dans le fichier: {cin}.",
                    value=cin,
                )
                continue
            seen_cins.add(cin)

        telephone, phone_valid = _normalize_phone(value(row, "Telephone"))
        if not phone_valid:
            failed += 1
            _add_issue(
                issues,
                row_number=row_number,
                column="Telephone",
                code="invalid_phone",
                message=f"Telephone invalide pour {matricule}.",
                value=value(row, "Telephone"),
            )
            continue

        employer_name = _safe_str(value(row, "Raison Sociale"))
        target_employer = default_employer
        if employer_name:
            found = employer_by_name.get(employer_name.lower())
            if not found:
                failed += 1
                _add_issue(
                    issues,
                    row_number=row_number,
                    column="Raison Sociale",
                    code="unknown_employer",
                    message=f"Employeur introuvable: {employer_name}.",
                    value=employer_name,
                )
                continue
            target_employer = found
        if not can_manage_worker(db, user, employer_id=target_employer.id):
            failed += 1
            _add_issue(
                issues,
                row_number=row_number,
                code="forbidden_employer_scope",
                message=f"Droits insuffisants sur l employeur {target_employer.raison_sociale}.",
            )
            continue

        regime_text = _safe_str(value(row, "Type Regime (Agricole/Non Agricole)")).lower()
        sector_text = _safe_str(value(row, "Secteur (agricole/non_agricole)")).lower()
        if "agricole" in f"{regime_text} {sector_text}" and "non" not in f"{regime_text} {sector_text}":
            selected_regime = next((item for item in regimes if item.code == "agricole"), default_regime)
        else:
            selected_regime = next((item for item in regimes if item.code == "non_agricole"), default_regime)
        date_debut_contrat, date_debut_valid = _parse_optional_date_strict(value(row, "Date Debut Contrat (JJ/MM/AAAA)"))
        date_embauche_alt, date_embauche_valid = _parse_optional_date_strict(value(row, "Date Embauche (JJ/MM/AAAA)"))
        if not date_debut_valid or not date_embauche_valid:
            failed += 1
            _add_issue(
                issues,
                row_number=row_number,
                column="Date Debut Contrat (JJ/MM/AAAA)" if not date_debut_valid else "Date Embauche (JJ/MM/AAAA)",
                code="invalid_date",
                message=f"Format de date invalide pour {matricule}. Utiliser JJ/MM/AAAA ou YYYY-MM-DD.",
            )
            continue
        date_embauche = date_debut_contrat or date_embauche_alt
        date_fin_contrat, date_fin_valid = _parse_optional_date_strict(value(row, "Date Fin Contrat (JJ/MM/AAAA)"))
        date_fin_essai, date_fin_essai_valid = _parse_optional_date_strict(value(row, "Date Fin Essai (JJ/MM/AAAA)"))
        date_naissance, date_naissance_valid = _parse_optional_date_strict(value(row, "Date de Naissance (JJ/MM/AAAA)"))
        cin_delivre_le, cin_delivre_le_valid = _parse_optional_date_strict(value(row, "CIN Delivre le (JJ/MM/AAAA)"))
        invalid_date_columns = []
        if not date_fin_valid:
            invalid_date_columns.append("Date Fin Contrat (JJ/MM/AAAA)")
        if not date_fin_essai_valid:
            invalid_date_columns.append("Date Fin Essai (JJ/MM/AAAA)")
        if not date_naissance_valid:
            invalid_date_columns.append("Date de Naissance (JJ/MM/AAAA)")
        if not cin_delivre_le_valid:
            invalid_date_columns.append("CIN Delivre le (JJ/MM/AAAA)")
        if invalid_date_columns:
            failed += 1
            _add_issue(
                issues,
                row_number=row_number,
                column=invalid_date_columns[0],
                code="invalid_date",
                message=f"Formats de date invalides pour {matricule}: {', '.join(invalid_date_columns)}.",
            )
            continue
        salary_base, salary_base_valid = _parse_optional_float_strict(value(row, "Salaire Base"), 0.0)
        horaire_hebdo, horaire_hebdo_valid = _parse_optional_float_strict(value(row, "Horaire Hebdo"), 40.0)
        taux_horaire_raw, taux_horaire_valid = _parse_optional_float_strict(value(row, "Taux Horaire"), 0.0)
        nb_enfants, nb_enfants_valid = _parse_optional_int_strict(value(row, "Nombre Enfants"), 0)
        duree_essai, duree_essai_valid = _parse_optional_int_strict(value(row, "Duree Essai (jours)"), 0)
        invalid_numeric_columns = []
        if not salary_base_valid:
            invalid_numeric_columns.append("Salaire Base")
        if not horaire_hebdo_valid:
            invalid_numeric_columns.append("Horaire Hebdo")
        if not taux_horaire_valid:
            invalid_numeric_columns.append("Taux Horaire")
        if not nb_enfants_valid:
            invalid_numeric_columns.append("Nombre Enfants")
        if not duree_essai_valid:
            invalid_numeric_columns.append("Duree Essai (jours)")
        if invalid_numeric_columns:
            failed += 1
            _add_issue(
                issues,
                row_number=row_number,
                column=invalid_numeric_columns[0],
                code="invalid_numeric",
                message=f"Valeurs numeriques invalides pour {matricule}: {', '.join(invalid_numeric_columns)}.",
            )
            continue
        if salary_base < 0 or horaire_hebdo < 0 or taux_horaire_raw < 0 or nb_enfants < 0 or duree_essai < 0:
            failed += 1
            _add_issue(
                issues,
                row_number=row_number,
                code="negative_numeric",
                message=f"Les valeurs numeriques negatives sont interdites pour {matricule}.",
            )
            continue
        vhm_value = selected_regime.vhm or (173.33 if horaire_hebdo <= 40 else 200.0)
        salary_horaire = taux_horaire_raw
        if salary_horaire <= 0 and vhm_value > 0:
            salary_horaire = salary_base / vhm_value if salary_base > 0 else 0.0

        payload = {
            "nom": nom,
            "prenom": _safe_str(value(row, "Prenom")),
            "sexe": _safe_str(value(row, "Sexe (M/F)")).upper(),
            "situation_familiale": _safe_str(value(row, "Situation Familiale")),
            "date_naissance": date_naissance,
            "lieu_naissance": _safe_str(value(row, "Lieu de Naissance")),
            "date_embauche": date_embauche,
            "nature_contrat": _safe_str(value(row, "Nature du Contrat")) or "CDI",
            "duree_essai_jours": duree_essai,
            "date_fin_essai": date_fin_essai,
            "date_debauche": date_fin_contrat,
            "salaire_base": salary_base,
            "salaire_horaire": salary_horaire,
            "horaire_hebdo": horaire_hebdo,
            "vhm": vhm_value,
            "etablissement": _safe_str(value(row, "Etablissement")),
            "departement": _safe_str(value(row, "Departement")),
            "service": _safe_str(value(row, "Service")),
            "unite": _safe_str(value(row, "Unite")),
            "adresse": _safe_str(value(row, "Adresse")),
            "telephone": telephone,
            "email": email,
            "cin": cin,
            "cin_delivre_le": cin_delivre_le,
            "cin_lieu": _safe_str(value(row, "CIN Lieu de delivrance")),
            "cnaps_num": _safe_str(value(row, "Numero CNaPS")),
            "nombre_enfant": nb_enfants,
            "poste": _safe_str(value(row, "Poste Actuel")),
            "categorie_prof": _safe_str(value(row, "Categorie Professionnelle")),
            "indice": _safe_str(value(row, "Indice Classification")),
            "mode_paiement": _safe_str(value(row, "Mode de Paiement")) or "Virement",
            "rib": _safe_str(value(row, "RIB")),
            "banque": _safe_str(value(row, "Nom de la Banque")),
            "nom_guichet": _safe_str(value(row, "Nom du Guichet")),
            "bic": _safe_str(value(row, "BIC / SWIFT")),
            "code_banque": _safe_str(value(row, "Code Banque")),
            "code_guichet": _safe_str(value(row, "Code Guichet")),
            "compte_num": _safe_str(value(row, "Numero de Compte")),
            "cle_rib": _safe_str(value(row, "Cle RIB")),
            "smie_agence": _safe_str(value(row, "SMIE Agence")),
            "smie_carte_num": _safe_str(value(row, "SMIE Numero Carte")),
            "avantage_vehicule": _safe_float(value(row, "Avantage Vehicule"), 0.0),
            "avantage_logement": _safe_float(value(row, "Avantage Logement"), 0.0),
            "avantage_telephone": _safe_float(value(row, "Avantage Telephone"), 0.0),
            "avantage_autres": _safe_float(value(row, "Avantage Autres"), 0.0),
            "solde_conge_initial": _safe_float(value(row, "Solde Conge Initial"), 0.0),
            "secteur": sector_text,
            "type_regime_id": selected_regime.id,
            "employer_id": target_employer.id,
        }

        candidate_summary = " | ".join(
            part for part in [_safe_str(value(row, "Observations")), _safe_str(value(row, "Autres Diplomes"))] if part
        )

        try:
            with db.begin_nested():
                existing = db.query(models.Worker).filter(models.Worker.matricule == matricule).first()
                if cin:
                    duplicate_cin_worker = (
                        db.query(models.Worker)
                        .filter(models.Worker.cin == cin, models.Worker.matricule != matricule)
                        .first()
                    )
                    if duplicate_cin_worker and (existing is None or duplicate_cin_worker.id != existing.id):
                        raise HTTPException(status_code=400, detail=f"CIN deja utilise par le matricule {duplicate_cin_worker.matricule}.")
                if email:
                    duplicate_email_worker = (
                        db.query(models.Worker)
                        .filter(func.lower(models.Worker.email) == email.lower(), models.Worker.matricule != matricule)
                        .first()
                    )
                    if duplicate_email_worker and (existing is None or duplicate_email_worker.id != existing.id):
                        raise HTTPException(status_code=400, detail=f"Email deja utilise par le matricule {duplicate_email_worker.matricule}.")
                if existing:
                    if not can_manage_worker(db, user, worker=existing):
                        raise HTTPException(status_code=403, detail=f"Droits insuffisants pour modifier {matricule}.")
                    if not update_existing:
                        skipped += 1
                        _add_issue(
                            issues,
                            row_number=row_number,
                            column="Matricule",
                            code="existing_skipped",
                            message=f"Matricule {matricule} deja existant (ignore).",
                            value=matricule,
                        )
                        continue
                    for field, field_value in payload.items():
                        if _has_value(field_value):
                            setattr(existing, field, field_value)
                    worker = existing
                    updated += 1
                else:
                    worker = models.Worker(matricule=matricule, **payload)
                    db.add(worker)
                    db.flush()
                    if worker.poste:
                        db.add(
                            models.WorkerPositionHistory(
                                worker_id=worker.id,
                                poste=worker.poste,
                                categorie_prof=worker.categorie_prof,
                                indice=worker.indice,
                                start_date=worker.date_embauche or datetime.now().date(),
                            )
                        )
                    created += 1

                candidate = _upsert_recruitment_candidate(
                    db=db,
                    employer_id=worker.employer_id,
                    first_name=_safe_str(value(row, "Prenom")),
                    last_name=nom.title(),
                    email=_safe_str(value(row, "Email")).lower(),
                    phone=_safe_str(value(row, "Telephone")),
                    education_level=_safe_str(value(row, "Niveau Etudes Principal")),
                    experience_years=_safe_float(value(row, "Annees Experience"), 0.0),
                    source_recruitment=_safe_str(value(row, "Source Recrutement")),
                    recruitment_type=_safe_str(value(row, "Type Recrutement")),
                    onboarding_status=_safe_str(value(row, "Statut Onboarding")),
                    candidate_reference=_safe_str(value(row, "Reference Candidat")),
                    summary=candidate_summary,
                )
                skill_names = _split_multi_values(value(row, "Competences (separees par ;)"))
                language_names = [f"Langue - {item}" for item in _split_multi_values(value(row, "Langues (separees par ;)"))]
                _upsert_worker_skills(db, worker, skill_names + language_names)
                sync_worker_master_data(db, worker, candidate=candidate)
                processed_rows += 1
        except HTTPException as exc:
            failed += 1
            logger.warning("workers.import row_rejected row=%s matricule=%s reason=%s", row_number, matricule, exc.detail)
            _add_issue(issues, row_number=row_number, code="business_rule_error", message=str(exc.detail))
        except Exception as exc:
            failed += 1
            logger.exception("workers.import row_failed row=%s matricule=%s", row_number, matricule)
            _add_issue(issues, row_number=row_number, code="unexpected_error", message=f"Erreur inattendue: {exc}")

    if dry_run:
        db.rollback()

    report = _build_report(
        mode=mode,
        total_rows=total_rows,
        processed_rows=processed_rows,
        created=created,
        updated=updated,
        skipped=skipped,
        failed=failed,
        unknown_columns=unknown_columns,
        missing_columns=missing_columns,
        issues=issues,
    )
    return report, created, updated, skipped


def _response_from_report(report: schemas.TabularImportReport) -> dict[str, Any]:
    return {
        "imported": report.created,
        "updated": report.updated,
        "skipped": report.skipped,
        "errors": [issue.message for issue in report.issues],
        "report": report.model_dump(mode="json"),
    }


@router.get("/template")
def get_workers_import_template(
    prefilled: bool = Query(False, description="Template pre-rempli avec salaries existants"),
    employer_id: Optional[int] = Query(None),
    export_format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    db: Session = Depends(get_db),
    user: models.AppUser = Depends(require_roles(*WRITE_RH_ROLES)),
):
    df = _build_template_dataframe(db, user, prefilled=prefilled, employer_id=employer_id)
    if export_format == "csv":
        content = dataframe_to_csv_bytes(df)
        filename = "modele_import_salaries.csv" if not prefilled else "salaries_existants.csv"
        return Response(content=content, media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"'})
    content = _build_workers_template_xlsx(df)
    filename = "modele_import_salaries.xlsx" if not prefilled else "salaries_existants.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/map-template")
def map_workers_import_template(
    file: UploadFile = File(...),
    user: models.AppUser = Depends(require_roles(*WRITE_RH_ROLES)),
):
    logger.info(
        "workers.import.map_template filename=%s user_id=%s",
        file.filename,
        getattr(user, "id", None),
    )
    try:
        mapped_df = map_user_excel_to_template(
            file.file.read(),
            WORKER_TEMPLATE_COLUMNS,
            aliases=COLUMN_ALIASES,
        )
    except DataMappingError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    content = _build_workers_template_xlsx(mapped_df)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="salaries_mapped_siirh.xlsx"'},
    )


@router.post("/preview", response_model=schemas.TabularImportReport)
def preview_workers_import(
    file: UploadFile = File(...),
    update_existing: bool = Form(False),
    db: Session = Depends(get_db),
    user: models.AppUser = Depends(require_roles(*WRITE_RH_ROLES)),
):
    logger.info(
        "workers.import.preview filename=%s update_existing=%s user_id=%s",
        file.filename,
        update_existing,
        getattr(user, "id", None),
    )
    content = file.file.read()
    df = read_tabular_bytes(content, file.filename)
    report, _, _, _ = _import_workers_dataframe(df=df, update_existing=update_existing, db=db, user=user, dry_run=True)
    logger.info(
        "workers.import.preview.completed filename=%s total_rows=%s created=%s updated=%s failed=%s skipped=%s",
        file.filename,
        report.total_rows,
        report.created,
        report.updated,
        report.failed,
        report.skipped,
    )
    return report


@router.post("")
def import_workers(
    file: UploadFile = File(...),
    update_existing: bool = Form(False),
    db: Session = Depends(get_db),
    user: models.AppUser = Depends(require_roles(*WRITE_RH_ROLES)),
):
    logger.info(
        "workers.import.start filename=%s update_existing=%s user_id=%s",
        file.filename,
        update_existing,
        getattr(user, "id", None),
    )
    content = file.file.read()
    df = read_tabular_bytes(content, file.filename)
    report, created, updated, skipped = _import_workers_dataframe(df=df, update_existing=update_existing, db=db, user=user, dry_run=False)
    if created > 0 or updated > 0:
        record_audit(
            db,
            actor=user,
            action="workers.import",
            entity_type="worker_import",
            entity_id=f"{created}:{updated}:{skipped}",
            route="/workers/import",
            after=report.model_dump(mode="json"),
        )
        db.commit()
    else:
        db.rollback()
    logger.info(
        "workers.import.completed filename=%s total_rows=%s created=%s updated=%s failed=%s skipped=%s",
        file.filename,
        report.total_rows,
        report.created,
        report.updated,
        report.failed,
        report.skipped,
    )
    return _response_from_report(report)
