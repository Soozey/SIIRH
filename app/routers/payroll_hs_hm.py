"""
Router pour gÃ©rer l'importation et la liaison des HS/HM aux paies
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from typing import List, Optional
import logging
import openpyxl
from io import BytesIO

from ..config.config import get_db
from ..models import PayrollHsHm, Worker, PayrollRun, HSCalculationHS, Absence, Avance, PayVar, Employer
from ..schemas import (
    PayrollHsHmCreate,
    PayrollHsHmOut,
    PayrollHsHmBase,
    LinkHsCalculationRequest,
    ExcelImportSummary
)
from ..utils.hs_hm_calculations import calculate_hs_hm_amounts
from fastapi.responses import StreamingResponse
from ..security import PAYROLL_WRITE_ROLES, READ_PAYROLL_ROLES, can_access_employer, require_roles
from ..services.organizational_filters import apply_worker_hierarchy_filters

router = APIRouter(prefix="/payroll-hs-hm", tags=["Payroll HS/HM"])
logger = logging.getLogger(__name__)


@router.get("/template")
def get_hs_hm_template(
    payroll_run_id: Optional[int] = None,
    employer_id: Optional[int] = None,
    etablissement: Optional[str] = Query(None),
    departement: Optional[str] = Query(None),
    service: Optional[str] = Query(None),
    unite: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user=Depends(require_roles(*READ_PAYROLL_ROLES)),
):
    """Telecharge le modele Excel pour l'import HS/HM."""

    payroll_run = None
    target_employer_id = employer_id

    if payroll_run_id:
        payroll_run = db.query(PayrollRun).filter(PayrollRun.id == payroll_run_id).first()
        if not payroll_run:
            raise HTTPException(status_code=404, detail="Payroll run not found")
        if not can_access_employer(db, user, payroll_run.employer_id):
            raise HTTPException(status_code=403, detail="Forbidden")
        target_employer_id = payroll_run.employer_id

    if target_employer_id is not None and not can_access_employer(db, user, target_employer_id):
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Paie"

        headers = [
            "Matricule",
            "Nom",
            "Prenom",
            "HSNI_130",
            "HSI_130",
            "HSNI_150",
            "HSI_150",
            "HMNH",
            "HMNO",
            "HMD",
            "HMJF",
            "ABSM_J",
            "ABSM_H",
            "ABSNR_J",
            "ABSNR_H",
            "ABSMP",
            "ABS1_J",
            "ABS1_H",
            "ABS2_J",
            "ABS2_H",
            "Avance",
            "Autre deduction 1",
            "Autre deduction 2",
            "Autre deduction 3",
            "Avantage Vehicule",
            "Avantage Logement",
            "Avantage Telephone",
            "Autres Avantages",
        ]
        ws.append(headers)

        from openpyxl.styles import Font, PatternFill, Alignment

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        if target_employer_id:
            workers = apply_worker_hierarchy_filters(
                db.query(Worker).filter(Worker.employer_id == target_employer_id),
                db,
                employer_id=target_employer_id,
                filters={
                    "etablissement": etablissement,
                    "departement": departement,
                    "service": service,
                    "unite": unite,
                },
            ).order_by(Worker.matricule, Worker.nom, Worker.prenom).all()

            for idx, worker in enumerate(workers, start=2):
                ws.cell(row=idx, column=1, value=worker.matricule or "")
                ws.cell(row=idx, column=2, value=worker.nom or "")
                ws.cell(row=idx, column=3, value=worker.prenom or "")

        column_widths = {
            "A": 16,
            "B": 20,
            "C": 24,
            "D": 12,
            "E": 12,
            "F": 12,
            "G": 12,
            "H": 12,
            "I": 12,
            "J": 12,
            "K": 12,
            "L": 12,
            "M": 12,
            "N": 12,
            "O": 12,
            "P": 12,
            "Q": 12,
            "R": 12,
            "S": 12,
            "T": 12,
            "U": 14,
            "V": 18,
            "W": 18,
            "X": 18,
            "Y": 18,
            "Z": 18,
            "AA": 18,
            "AB": 18,
        }
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "Modele_Import_Paie.xlsx"
        if target_employer_id:
            employer = db.query(Employer).filter(Employer.id == target_employer_id).first()
            emp_name = getattr(employer, "raison_sociale", f"EMP{target_employer_id}") if employer else f"EMP{target_employer_id}"
            safe_emp_name = "".join([c for c in emp_name if c.isalnum() or c in (" ", "-", "_")]).strip()
            filename = f"Import_Paie_{safe_emp_name}.xlsx"
            if payroll_run:
                filename = f"Import_Paie_{safe_emp_name}_{payroll_run.period}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        logger.exception("Erreur lors de la generation du modele d'import paie")
        raise HTTPException(status_code=500, detail=f"Error generating template: {str(e)}")


@router.post("/{payroll_run_id}/link-manual/{worker_id}", response_model=PayrollHsHmOut)
def link_manual_hs_calculation(
    payroll_run_id: int,
    worker_id: int,
    request: LinkHsCalculationRequest,
    db: Session = Depends(get_db),
    user=Depends(require_roles(*PAYROLL_WRITE_ROLES)),
):
    """
    Lie un calcul HS manuel (from HeuresSupplementairesPageHS) Ã  une paie.
    RÃ©cupÃ¨re les heures depuis hs_calculations_HS et calcule les montants.
    """
    # Verify payroll run exists
    payroll_run = db.query(PayrollRun).filter(PayrollRun.id == payroll_run_id).first()
    if not payroll_run:
        raise HTTPException(status_code=404, detail="Payroll run not found")
    if not can_access_employer(db, user, payroll_run.employer_id):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    # Verify worker exists
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    if worker.employer_id != payroll_run.employer_id:
        raise HTTPException(status_code=400, detail="Worker/employer mismatch for payroll run")
    
    # Verify HS calculation exists
    hs_calc = db.query(HSCalculationHS).filter(
        HSCalculationHS.id_HS == request.hs_calculation_id
    ).first()
    if not hs_calc:
        raise HTTPException(status_code=404, detail="HS calculation not found")
    
    # Check if HS/HM already exists for this payroll + worker
    existing = db.query(PayrollHsHm).filter(
        PayrollHsHm.payroll_run_id == payroll_run_id,
        PayrollHsHm.worker_id == worker_id
    ).first()
    
    if existing:
        # Update existing
        db.delete(existing)
        db.commit()
    
    # Prepare hours dict
    hours = {
        'hsni_130_heures': float(hs_calc.total_HSNI_130_heures_HS),
        'hsi_130_heures': float(hs_calc.total_HSI_130_heures_HS),
        'hsni_150_heures': float(hs_calc.total_HSNI_150_heures_HS),
        'hsi_150_heures': float(hs_calc.total_HSI_150_heures_HS),
        'hmnh_heures': float(hs_calc.total_HMNH_30_heures_HS),
        'hmno_heures': float(hs_calc.total_HMNO_50_heures_HS),
        'hmd_heures': float(hs_calc.total_HMD_40_heures_HS),
        'hmjf_heures': float(hs_calc.total_HMJF_50_heures_HS),
    }
    
    # Calculate amounts in Ariary
    amounts = calculate_hs_hm_amounts(hours, worker.salaire_horaire)
    
    # Create new PayrollHsHm
    payroll_hs_hm = PayrollHsHm(
        payroll_run_id=payroll_run_id,
        worker_id=worker_id,
        source_type="MANUAL",
        hs_calculation_id=hs_calc.id_HS,
        **hours,
        **amounts
    )
    
    db.add(payroll_hs_hm)
    db.commit()
    db.refresh(payroll_hs_hm)
    
    return payroll_hs_hm


@router.post("/{payroll_run_id}/import-excel", response_model=ExcelImportSummary)
async def import_hs_hm_from_excel(
    payroll_run_id: int,
    file: UploadFile = File(...),
    etablissement: Optional[str] = Query(None),
    departement: Optional[str] = Query(None),
    service: Optional[str] = Query(None),
    unite: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user=Depends(require_roles(*PAYROLL_WRITE_ROLES)),
):
    """
    Import HS/HM, Absences, and Avances data from Excel file.
    Expected format: Matricule, HSNI_130, HSI_130, HSNI_150, HSI_150, HMNH, HMNO, HMD, HMJF,
                     ABSM_J, ABSM_H, ABSNR_J, ABSNR_H, ABSMP, ABS1_J, ABS1_H, ABS2_J, ABS2_H, Avance
    """
    # Verify payroll run exists
    payroll_run = db.query(PayrollRun).filter(PayrollRun.id == payroll_run_id).first()
    if not payroll_run:
        raise HTTPException(status_code=404, detail="Payroll run not found")
    if not can_access_employer(db, user, payroll_run.employer_id):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    # Read Excel file
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")
    
    total_rows = 0
    successful = 0
    failed = 0
    errors = []
    filter_active = any([etablissement, departement, service, unite])
    allowed_worker_ids = {
        worker.id
        for worker in apply_worker_hierarchy_filters(
            db.query(Worker).filter(Worker.employer_id == payroll_run.employer_id),
            db,
            employer_id=payroll_run.employer_id,
            filters={
                "etablissement": etablissement,
                "departement": departement,
                "service": service,
                "unite": unite,
            },
        ).all()
    }

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    normalized_headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]
    has_identity_columns = len(normalized_headers) >= 3 and normalized_headers[:3] == ["matricule", "nom", "prenom"]
    data_start_index = 3 if has_identity_columns else 1

    # Skip header row, start from row 2
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue
        
        total_rows += 1
        
        try:
            # Parse row
            matricule = str(row[0]).strip()
            
            # Find worker by matricule
            worker = db.query(Worker).filter(
                Worker.matricule == matricule,
                Worker.employer_id == payroll_run.employer_id,
            ).first()
            if not worker:
                errors.append(f"Row {row_idx}: Worker '{matricule}' not found")
                failed += 1
                continue
            if filter_active and worker.id not in allowed_worker_ids:
                errors.append(f"Row {row_idx}: Worker '{matricule}' hors du filtre organisationnel actif")
                failed += 1
                continue
            
            def read_number(index: int) -> float:
                if len(row) <= index:
                    return 0.0
                value = row[index]
                if value in (None, ""):
                    return 0.0
                return float(value)

            # Parse hours (default 0.0 if empty)
            hours = {
                'hsni_130_heures': read_number(data_start_index + 0),
                'hsi_130_heures': read_number(data_start_index + 1),
                'hsni_150_heures': read_number(data_start_index + 2),
                'hsi_150_heures': read_number(data_start_index + 3),
                'hmnh_heures': read_number(data_start_index + 4),
                'hmno_heures': read_number(data_start_index + 5),
                'hmd_heures': read_number(data_start_index + 6),
                'hmjf_heures': read_number(data_start_index + 7),
            }
            
            # Parse absences (colonnes 9-17)
            absences_data = {
                'ABSM_J': read_number(data_start_index + 8),
                'ABSM_H': read_number(data_start_index + 9),
                'ABSNR_J': read_number(data_start_index + 10),
                'ABSNR_H': read_number(data_start_index + 11),
                'ABSMP': read_number(data_start_index + 12),
                'ABS1_J': read_number(data_start_index + 13),
                'ABS1_H': read_number(data_start_index + 14),
                'ABS2_J': read_number(data_start_index + 15),
                'ABS2_H': read_number(data_start_index + 16),
            }
            
            # Parse avance (colonne 18)
            avance_montant = read_number(data_start_index + 17)
            
            # Validate all hours >= 0
            if any(h < 0 for h in hours.values()):
                errors.append(f"Row {row_idx}: Negative hours not allowed")
                failed += 1
                continue
            
            # Validate absences >= 0
            if any(v < 0 for v in absences_data.values()):
                errors.append(f"Row {row_idx}: Negative absence values not allowed")
                failed += 1
                continue
            
            # Calculate HS/HM amounts
            amounts = calculate_hs_hm_amounts(hours, worker.salaire_horaire)
            
            # === HS/HM ===
            # Check if already exists
            existing = db.query(PayrollHsHm).filter(
                PayrollHsHm.payroll_run_id == payroll_run_id,
                PayrollHsHm.worker_id == worker.id
            ).first()
            
            if existing:
                # Update
                for key, value in hours.items():
                    setattr(existing, key, value)
                for key, value in amounts.items():
                    setattr(existing, key, value)
                existing.source_type = "IMPORT"
                existing.import_file_name = file.filename
                existing.hs_calculation_id = None
            else:
                # Create new
                payroll_hs_hm = PayrollHsHm(
                    payroll_run_id=payroll_run_id,
                    worker_id=worker.id,
                    source_type="IMPORT",
                    import_file_name=file.filename,
                    **hours,
                    **amounts
                )
                db.add(payroll_hs_hm)
            
            # === ABSENCES ===
            # Get period from payroll_run
            period = payroll_run.period  # Format YYYY-MM
            
            existing_absence = db.query(Absence).filter(
                Absence.worker_id == worker.id,
                Absence.mois == period
            ).first()
            
            if existing_absence:
                # Update
                for key, value in absences_data.items():
                    setattr(existing_absence, key, value)
            else:
                # Create new
                absence = Absence(
                    worker_id=worker.id,
                    mois=period,
                    **absences_data
                )
                db.add(absence)
            
            # === AVANCES ===
            if avance_montant > 0:
                existing_avance = db.query(Avance).filter(
                    Avance.worker_id == worker.id,
                    Avance.periode == period
                ).first()
                
                if existing_avance:
                    existing_avance.montant = avance_montant
                else:
                    avance = Avance(
                        worker_id=worker.id,
                        periode=period,
                        montant=avance_montant
                    )
                    db.add(avance)

            # === PAYVAR (Autres Déductions & Avantages Override) ===
            # Columns 19..25 (0-indexed in row): 19=AutreDed1 ... 25=AutreAv
            # Row indices: 
            # 19: Autre Ded 1
            # 20: Autre Ded 2
            # 21: Autre Ded 3
            # 22: Avantage Véhicule
            # 23: Avantage Logement
            # 24: Avantage Téléphone
            # 25: Autres Avantages

            autre_ded1 = read_number(data_start_index + 18)
            autre_ded2 = read_number(data_start_index + 19)
            autre_ded3 = read_number(data_start_index + 20)
            
            av_vehicule = read_number(data_start_index + 21)
            av_logement = read_number(data_start_index + 22)
            av_telephone = read_number(data_start_index + 23)
            av_autres = read_number(data_start_index + 24)

            # Check if we have ANY data to save in PayVar
            if any([autre_ded1, autre_ded2, autre_ded3, av_vehicule, av_logement, av_telephone, av_autres]):
                # Get or Create PayVar
                payvar_entry = db.query(PayVar).filter(
                    PayVar.worker_id == worker.id,
                    PayVar.period == period
                ).first()

                if not payvar_entry:
                    payvar_entry = PayVar(worker_id=worker.id, period=period)
                    db.add(payvar_entry)
                
                # Update fields
                payvar_entry.autre_ded1 = autre_ded1
                payvar_entry.autre_ded2 = autre_ded2
                payvar_entry.autre_ded3 = autre_ded3
                payvar_entry.avantage_vehicule = av_vehicule
                payvar_entry.avantage_logement = av_logement
                payvar_entry.avantage_telephone = av_telephone
                payvar_entry.avantage_autres = av_autres

            successful += 1
            
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
            failed += 1
            continue
    
    db.commit()
    
    return ExcelImportSummary(
        total_rows=total_rows,
        successful=successful,
        failed=failed,
        errors=errors
    )


@router.get("/{payroll_run_id}/all", response_model=List[PayrollHsHmOut])
def get_all_hs_hm_for_payroll(
    payroll_run_id: int,
    etablissement: Optional[str] = Query(None),
    departement: Optional[str] = Query(None),
    service: Optional[str] = Query(None),
    unite: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user=Depends(require_roles(*READ_PAYROLL_ROLES)),
):
    """Get all HS/HM data for a payroll run, merged with Absences"""
    
    # 1. Get PayrollRun to determine period
    payroll_run = db.query(PayrollRun).filter(PayrollRun.id == payroll_run_id).first()
    if not payroll_run:
        return []
    if not can_access_employer(db, user, payroll_run.employer_id):
        raise HTTPException(status_code=403, detail="Forbidden")
        
    period = payroll_run.period
    scoped_workers = apply_worker_hierarchy_filters(
        db.query(Worker).filter(Worker.employer_id == payroll_run.employer_id),
        db,
        employer_id=payroll_run.employer_id,
        filters={
            "etablissement": etablissement,
            "departement": departement,
            "service": service,
            "unite": unite,
        },
    ).all()
    scoped_worker_ids = {worker.id for worker in scoped_workers}
    if not scoped_worker_ids:
        return []
    
    # 2. Get all HS/HM
    hs_hms = db.query(PayrollHsHm).filter(
        PayrollHsHm.payroll_run_id == payroll_run_id,
        PayrollHsHm.worker_id.in_(scoped_worker_ids),
    ).all()
    
    # 3. Get all Absences
    absences = db.query(Absence).filter(
        Absence.mois == period,
        Absence.worker_id.in_(scoped_worker_ids),
    ).all()
    
    # 4. Get all Avances
    avances = db.query(Avance).filter(
        Avance.periode == period,
        Avance.worker_id.in_(scoped_worker_ids),
    ).all()

    # 5. Get all PayVars (Advantages & Deductions)
    payvars = db.query(PayVar).filter(
        PayVar.period == period,
        PayVar.worker_id.in_(scoped_worker_ids),
    ).all()
    
    # Merge Data
    hs_map = {hs.worker_id: hs for hs in hs_hms}
    abs_map = {a.worker_id: a for a in absences}
    av_map = {a.worker_id: a for a in avances}
    pv_map = {p.worker_id: p for p in payvars}
    
    all_worker_ids = set(hs_map.keys()) | set(abs_map.keys()) | set(av_map.keys()) | set(pv_map.keys())
    
    result = []
    
    for wid in all_worker_ids:
        hs_entry = hs_map.get(wid)
        abs_entry = abs_map.get(wid)
        av_entry = av_map.get(wid)
        pv_entry = pv_map.get(wid)
        
        # Base dict
        if hs_entry:
            out_dict = {c.name: getattr(hs_entry, c.name) for c in PayrollHsHm.__table__.columns}
        else:
            out_dict = {
                "id": None, # Should be None for new entries on frontend logic usually, check types? schema says int, let's put 0
                "payroll_run_id": payroll_run_id,
                "worker_id": wid,
                "source_type": "MANUAL",
                "hs_calculation_id": None,
                "import_file_name": None,
                "hsni_130_heures": 0.0, "hsi_130_heures": 0.0,
                "hsni_150_heures": 0.0, "hsi_150_heures": 0.0,
                "hmnh_heures": 0.0, "hmno_heures": 0.0,
                "hmd_heures": 0.0, "hmjf_heures": 0.0,
                "hsni_130_montant": 0.0, "hsi_130_montant": 0.0,
                "hsni_150_montant": 0.0, "hsi_150_montant": 0.0,
                "hmnh_montant": 0.0, "hmno_montant": 0.0,
                "hmd_montant": 0.0, "hmjf_montant": 0.0,
                "created_at": None, "updated_at": None,
                "id": 0 # Default ID 0
            }

        # Inject Absence Data
        if abs_entry:
            out_dict["ABSM_J"] = abs_entry.ABSM_J
            out_dict["ABSM_H"] = abs_entry.ABSM_H
            out_dict["ABSNR_J"] = abs_entry.ABSNR_J
            out_dict["ABSNR_H"] = abs_entry.ABSNR_H
            out_dict["ABSMP"] = abs_entry.ABSMP
            out_dict["ABS1_J"] = abs_entry.ABS1_J
            out_dict["ABS1_H"] = abs_entry.ABS1_H
            out_dict["ABS2_J"] = abs_entry.ABS2_J
            out_dict["ABS2_H"] = abs_entry.ABS2_H
        else:
            out_dict["ABSM_J"] = 0.0
            out_dict["ABSM_H"] = 0.0
            out_dict["ABSNR_J"] = 0.0
            out_dict["ABSNR_H"] = 0.0
            out_dict["ABSMP"] = 0.0
            out_dict["ABS1_J"] = 0.0
            out_dict["ABS1_H"] = 0.0
            out_dict["ABS2_J"] = 0.0
            out_dict["ABS2_H"] = 0.0
            
        # Inject Avance Data
        if av_entry:
            out_dict["avance"] = av_entry.montant
        else:
            out_dict["avance"] = 0.0

        # Inject PayVar Data (Advantages & Deductions)
        if pv_entry:
            out_dict["autre_ded1"] = pv_entry.autre_ded1
            out_dict["autre_ded2"] = pv_entry.autre_ded2
            out_dict["autre_ded3"] = pv_entry.autre_ded3
            out_dict["autre_ded4"] = pv_entry.autre_ded4
            out_dict["avantage_vehicule"] = pv_entry.avantage_vehicule
            out_dict["avantage_logement"] = pv_entry.avantage_logement
            out_dict["avantage_telephone"] = pv_entry.avantage_telephone
            out_dict["avantage_autres"] = pv_entry.avantage_autres
        else:
            out_dict["autre_ded1"] = 0.0
            out_dict["autre_ded2"] = 0.0
            out_dict["autre_ded3"] = 0.0
            out_dict["autre_ded4"] = 0.0
            out_dict["avantage_vehicule"] = 0.0
            out_dict["avantage_logement"] = 0.0
            out_dict["avantage_telephone"] = 0.0
            out_dict["avantage_autres"] = 0.0
            
        # Primes Data Removed from Schema

            
        result.append(out_dict)
        
    return result


@router.put("/{payroll_run_id}/{worker_id}", response_model=PayrollHsHmOut)
def update_worker_hs_hm(
    payroll_run_id: int,
    worker_id: int,
    payload: PayrollHsHmBase,
    db: Session = Depends(get_db),
    user=Depends(require_roles(*PAYROLL_WRITE_ROLES)),
):
    """
    Update HS/HM AND Absences for a specific worker.
    """
    # 1. Verify Worker & PayrollRun
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
        
    payroll_run = db.query(PayrollRun).filter(PayrollRun.id == payroll_run_id).first()
    if not payroll_run:
        raise HTTPException(status_code=404, detail="Payroll run not found")
    if not can_access_employer(db, user, payroll_run.employer_id):
        raise HTTPException(status_code=403, detail="Forbidden")
    if worker.employer_id != payroll_run.employer_id:
        raise HTTPException(status_code=400, detail="Worker/employer mismatch for payroll run")
    
    period = payroll_run.period

    # 2. Separate Payload
    data = payload.dict()
    
    hs_keys = [
        "hsni_130_heures", "hsi_130_heures",
        "hsni_150_heures", "hsi_150_heures",
        "hmnh_heures", "hmno_heures",
        "hmd_heures", "hmjf_heures"
    ]
    abs_keys = [
        "ABSM_J", "ABSM_H", 
        "ABSNR_J", "ABSNR_H", 
        "ABSMP", 
        "ABS1_J", "ABS1_H", 
        "ABS2_J", "ABS2_H"
    ]

    # prime_keys removed
    
    hs_data = {k: v for k, v in data.items() if k in hs_keys}
    abs_data = {k: v for k, v in data.items() if k in abs_keys}
    avance_val = data.get("avance", 0.0)

    payvar_keys = [
        "autre_ded1", "autre_ded2", "autre_ded3", "autre_ded4",
        "avantage_vehicule", "avantage_logement", "avantage_telephone", "avantage_autres"
    ]
    pv_data = {k: v for k, v in data.items() if k in payvar_keys}
    
    # 3. Update/Create HS/HM (PayrollHsHm)
    amounts = calculate_hs_hm_amounts(hs_data, worker.salaire_horaire)
    
    hs_hm = db.query(PayrollHsHm).filter(
        PayrollHsHm.payroll_run_id == payroll_run_id,
        PayrollHsHm.worker_id == worker_id
    ).first()
    
    if hs_hm:
        for as_key, as_val in hs_data.items():
            setattr(hs_hm, as_key, as_val)
        for amt_key, amt_val in amounts.items():
            setattr(hs_hm, amt_key, amt_val)
        hs_hm.source_type = "MANUAL"
    else:
        hs_hm = PayrollHsHm(
            payroll_run_id=payroll_run_id,
            worker_id=worker_id,
            source_type="MANUAL",
            **hs_data,
            **amounts
        )
        db.add(hs_hm)
    
    # 4. Update/Create Absence
    absence = db.query(Absence).filter(
        Absence.worker_id == worker_id,
        Absence.mois == period
    ).first()
    
    if absence:
        for k, v in abs_data.items():
            setattr(absence, k, v)
    else:
        absence = Absence(
            worker_id=worker_id,
            mois=period,
            **abs_data
        )
        db.add(absence)

    # 5. Update/Create Avance
    avance_obj = db.query(Avance).filter(
        Avance.worker_id == worker_id,
        Avance.periode == period
    ).first()
    
    if avance_obj:
        avance_obj.montant = avance_val
    elif avance_val > 0:
        # Only create if > 0
        avance_obj = Avance(
            worker_id=worker_id,
            periode=period,
            montant=avance_val
        )
        db.add(avance_obj)

    # 6. Update PayVar (Advantages & Deductions)
    payvar_entry = db.query(PayVar).filter(
        PayVar.worker_id == worker_id,
        PayVar.period == period
    ).first()

    should_create_pv = any(v != 0 for v in pv_data.values())
    
    if not payvar_entry and should_create_pv:
        payvar_entry = PayVar(worker_id=worker_id, period=period)
        db.add(payvar_entry)
        
    if payvar_entry:
        for k, v in pv_data.items():
            setattr(payvar_entry, k, v)


    db.commit()
    db.refresh(hs_hm)
    
    # 6. Build Response
    out_dict = {c.name: getattr(hs_hm, c.name) for c in PayrollHsHm.__table__.columns}
    
    for k in abs_keys:
        out_dict[k] = abs_data.get(k, 0.0)
    
    out_dict["avance"] = avance_val

    for k in payvar_keys:
        out_dict[k] = pv_data.get(k, 0.0)


    return out_dict 

@router.delete("/{payroll_run_id}/{worker_id}", status_code=204)
def delete_worker_hs_hm(
    payroll_run_id: int,
    worker_id: int,
    db: Session = Depends(get_db),
    user=Depends(require_roles(*PAYROLL_WRITE_ROLES)),
):
    """
    Supprime les HS/HM et les Absences pour un salariÃ© et une paie donnÃ©e.
    Remet effectivement Ã  zÃ©ro les compteurs.
    """
    # 1. Get PayrollRun to find period
    payroll_run = db.query(PayrollRun).filter(PayrollRun.id == payroll_run_id).first()
    if not payroll_run:
        raise HTTPException(status_code=404, detail="Payroll run not found")
    if not can_access_employer(db, user, payroll_run.employer_id):
        raise HTTPException(status_code=403, detail="Forbidden")
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if worker and worker.employer_id != payroll_run.employer_id:
        raise HTTPException(status_code=400, detail="Worker/employer mismatch for payroll run")
        
    period = payroll_run.period

    # 2. Delete HS/HM
    hs_hm = db.query(PayrollHsHm).filter(
        PayrollHsHm.payroll_run_id == payroll_run_id,
        PayrollHsHm.worker_id == worker_id
    ).first()
    
    if hs_hm:
        db.delete(hs_hm)
        
    # 3. Delete Absences (same period)
    absence = db.query(Absence).filter(
        Absence.worker_id == worker_id,
        Absence.mois == period
    ).first()
    
    if absence:
        db.delete(absence)
        
    # 4. Delete Avance
    avance = db.query(Avance).filter(
        Avance.worker_id == worker_id,
        Avance.periode == period
    ).first()
    
    if avance:
        db.delete(avance)

    # 5. Delete PayVar (Advantages & Deductions)
    payvar = db.query(PayVar).filter(
        PayVar.worker_id == worker_id,
        PayVar.period == period
    ).first()
    
    if payvar:
        db.delete(payvar)
        
    db.commit()
    return None

@router.post("/{payroll_run_id}/reset-bulk", status_code=204)
def reset_bulk_hs_hm(
    payroll_run_id: int,
    worker_ids: List[int],
    db: Session = Depends(get_db),
    user=Depends(require_roles(*PAYROLL_WRITE_ROLES)),
):
    """
    Supprime HS/HM et Absences pour une liste de salariÃ©s (Bulk Reset).
    """
    # 1. Get PayrollRun
    payroll_run = db.query(PayrollRun).filter(PayrollRun.id == payroll_run_id).first()
    if not payroll_run:
        raise HTTPException(status_code=404, detail="Payroll run not found")
    if not can_access_employer(db, user, payroll_run.employer_id):
        raise HTTPException(status_code=403, detail="Forbidden")
    if worker_ids:
        allowed_rows = db.query(Worker.id).filter(
            Worker.id.in_(worker_ids),
            Worker.employer_id == payroll_run.employer_id,
        ).all()
        allowed_ids = {row[0] for row in allowed_rows}
        if len(allowed_ids) != len(set(worker_ids)):
            raise HTTPException(status_code=400, detail="Some workers are outside payroll run employer scope")
        
    period = payroll_run.period
    
    # 2. Bulk Delete PayrollHsHm
    db.query(PayrollHsHm).filter(
        PayrollHsHm.payroll_run_id == payroll_run_id,
        PayrollHsHm.worker_id.in_(worker_ids)
    ).delete(synchronize_session=False)
    
    # 3. Bulk Delete Absences
    db.query(Absence).filter(
        Absence.mois == period,
        Absence.worker_id.in_(worker_ids)
    ).delete(synchronize_session=False)

    # 4. Bulk Delete Avances
    db.query(Avance).filter(
        Avance.periode == period,
        Avance.worker_id.in_(worker_ids)
    ).delete(synchronize_session=False)

    # 5. Bulk Delete PayVar
    db.query(PayVar).filter(
        PayVar.period == period,
        PayVar.worker_id.in_(worker_ids)
    ).delete(synchronize_session=False)
    
    db.commit()
    return None
