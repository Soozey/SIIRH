from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Form, Query
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List
import openpyxl
from io import BytesIO
from datetime import datetime

from ..config.config import get_db
from .. import models

router = APIRouter(
    prefix="/primes",
    tags=["primes"],
    responses={404: {"description": "Not found"}},
)

from pydantic import BaseModel
from typing import Optional

class PrimeBase(BaseModel):
    label: str
    description: Optional[str] = None
    formula_nombre: Optional[str] = None
    formula_base: Optional[str] = None
    formula_taux: Optional[str] = None
    operation_1: Optional[str] = "*"
    operation_2: Optional[str] = "*"
    is_active: bool = True
    is_cotisable: bool = True
    is_imposable: bool = True

class PrimeCreate(PrimeBase):
    employer_id: int

class PrimeOut(PrimeBase):
    id: int
    employer_id: int
    class Config:
        from_attributes = True

@router.get("/", response_model=List[PrimeOut])
def get_primes(employer_id: int, db: Session = Depends(get_db)):
    return db.query(models.Prime).filter(models.Prime.employer_id == employer_id).all()

@router.post("/", response_model=PrimeOut)
def create_prime(prime: PrimeCreate, db: Session = Depends(get_db)):
    db_prime = models.Prime(**prime.dict())
    db.add(db_prime)
    db.commit()
    db.refresh(db_prime)
    return db_prime

@router.put("/{prime_id}", response_model=PrimeOut)
def update_prime(prime_id: int, prime: PrimeCreate, db: Session = Depends(get_db)):
    db_prime = db.query(models.Prime).get(prime_id)
    if not db_prime:
        raise HTTPException(404, "Prime not found")
    
    for key, value in prime.dict().items():
        setattr(db_prime, key, value)
    
    db.commit()
    db.refresh(db_prime)
    return db_prime

@router.delete("/{prime_id}")
def delete_prime(prime_id: int, db: Session = Depends(get_db)):
    db_prime = db.query(models.Prime).get(prime_id)
    if not db_prime:
        raise HTTPException(404, "Prime not found")
    db.delete(db_prime)
    db.commit()
    return {"message": "Deleted"}

class AssociationRequest(BaseModel):
    worker_id: int
    prime_id: int
    is_active: bool = True

@router.post("/associations")
def set_worker_prime_association(assoc: AssociationRequest, db: Session = Depends(get_db)):
    link = db.query(models.WorkerPrimeLink).filter(
        models.WorkerPrimeLink.worker_id == assoc.worker_id,
        models.WorkerPrimeLink.prime_id == assoc.prime_id
    ).first()
    
    if not link:
        link = models.WorkerPrimeLink(worker_id=assoc.worker_id, prime_id=assoc.prime_id)
        db.add(link)
    
    link.is_active = assoc.is_active
    db.commit()
    return {"message": "Updated"}

@router.get("/associations", response_model=List[dict])
def get_worker_associations(worker_id: int, db: Session = Depends(get_db)):
    links = db.query(models.WorkerPrimeLink).filter(models.WorkerPrimeLink.worker_id == worker_id).all()
    # Return list of active IDs or detailed objects
    return [{"prime_id": l.prime_id, "is_active": l.is_active} for l in links]


@router.post("/reset-overrides")
def reset_prime_overrides(
    period: str = Query(...),
    employer_id: int = Query(...),
    db: Session = Depends(get_db)
):
    """
    Réinitialise manuellement les overrides de primes (PayrollPrime) pour une période donnée.
    Supprime toutes les valeurs importées pour revenir aux formules par défaut.
    """
    # Récupérer tous les workers de l'employeur
    workers = db.query(models.Worker).filter(models.Worker.employer_id == employer_id).all()
    worker_ids = [w.id for w in workers]
    
    if not worker_ids:
        return {"message": "Aucun salarié trouvé pour cet employeur", "deleted_count": 0}
    
    # Supprimer tous les overrides pour cette période
    deleted_count = db.query(models.PayrollPrime).filter(
        models.PayrollPrime.worker_id.in_(worker_ids),
        models.PayrollPrime.period == period
    ).delete(synchronize_session=False)
    
    db.commit()
    
    return {
        "message": "Les valeurs par défaut sont restaurées",
        "deleted_count": deleted_count
    }


@router.get("/template")
def download_primes_template(
    employer_id: int = Query(...),
    db: Session = Depends(get_db)
):
    """
    Génère un modèle Excel dynamique incluant Primes 1-5 (Nombre/Base), 13ème mois et Primes détaillées.
    """
    employer = db.query(models.Employer).filter(models.Employer.id == employer_id).first()
    if not employer:
        raise HTTPException(status_code=404, detail="Employer not found")

    # 1. Primes détaillées (Global Primes from PrimesManagement)
    # Use models.Prime instead of WorkerPrime to sync with "Gestion des Primes"
    primes = db.query(models.Prime).filter(models.Prime.employer_id == employer_id).all()
    labels = sorted([p.label for p in primes])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Primes"

    # En-têtes fixes
    headers = ["Matricule", "Nom", "Prénom"]
    
    # Primes Détaillées (Nombre / Base) - Purely Dynamic now
    for label in labels:
        headers.append(f"{label} (Nombre)")
        headers.append(f"{label} (Base)")

    ws.append(headers)

    # Pre-fill workers (filtered by Employer!)
    workers = db.query(models.Worker).filter(models.Worker.employer_id == employer_id).all()
    
    total_cols = (len(labels) * 2)
    
    for w in workers:
        row = [w.matricule, w.nom, w.prenom]
        row.extend([None] * total_cols)
        ws.append(row)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Modele_Import_Primes.xlsx"}
    )


@router.post("/import")
async def import_primes(
    file: UploadFile = File(...),
    period: str = Form(...),
    employer_id: int = Form(...),
    db: Session = Depends(get_db)
):
    """
    Importe les valeurs variables (Nombre/Base) des primes depuis un fichier Excel.
    Gère aussi les Primes 1-5 et 13ème Mois (PayVar).
    """
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Fichier invalide. Format attendu : .xlsx")

    # Fetch Employer for labels
    employer = db.query(models.Employer).filter(models.Employer.id == employer_id).first()
    if not employer:
        raise HTTPException(status_code=404, detail="Employer not found")

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active
        
        # 1. Analyser les en-têtes
        headers = [cell.value for cell in ws[1]]
        
        # Mapping index -> info
        col_mapping = {}

        # Prepare PayVar labels map
        pv_labels_map = {}
        
        for idx, header_val in enumerate(headers):
            if not header_val or not isinstance(header_val, str):
                continue
            header = header_val.strip()
            
            if header in ["Matricule", "Nom", "Prénom"]:
                continue
            
            # Check PayVar (Exact Match)
            if header in pv_labels_map:
                col_mapping[idx] = {"target": "payvar", "field": pv_labels_map[header]}
                continue
            
            # Check PayrollPrime (Suffix)
            if header.endswith(" (Nombre)"):
                label = header[:-9]  # Enlever " (Nombre)"
                col_mapping[idx] = {"target": "payroll_prime", "label": label, "type": "nombre"}
            elif header.endswith(" (Base)"):
                label = header[:-7]  # Enlever " (Base)"
                col_mapping[idx] = {"target": "payroll_prime", "label": label, "type": "base"}
            else:
                continue

        # 2. Parcourir les lignes
        count_updated = 0
        errors = []

        # Prepare PayVar clearing map once (outside loop)
        label_to_field = {
            '13ème Mois': 'prime_13',
            employer.label_prime1 or 'Prime 1': 'prime1',
            employer.label_prime2 or 'Prime 2': 'prime2',
            employer.label_prime3 or 'Prime 3': 'prime3',
            employer.label_prime4 or 'Prime 4': 'prime4',
            employer.label_prime5 or 'Prime 5': 'prime5',
        }

        # Itérer à partir de la ligne 2
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            matricule = row[0]
            if not matricule:
                continue # Ligne vide ou fin de fichier
            
            # Trouver le travailleur
            worker = db.query(models.Worker).filter(models.Worker.matricule == str(matricule)).first()
            if not worker:
                errors.append(f"Ligne {row_idx}: Matricule {matricule} inconnu.")
                continue
            
            # Cache PayVar fetch per worker
            pv = None
            
            # Cache existing PayrollPrimes for this worker/period to avoid Duplicate DB/Flush issues
            existing_primes = db.query(models.PayrollPrime).filter(
                models.PayrollPrime.worker_id == worker.id,
                models.PayrollPrime.period == period
            ).all()
            primes_map = {p.prime_label: p for p in existing_primes}

            # Traiter chaque colonne
            for col_idx, info in col_mapping.items():
                if col_idx >= len(row): continue
                val = row[col_idx]
                
                # Conversion sécurisée ou None
                val_float = None
                if val is not None and val != "":
                    try:
                        # Gérer les pourcentages Excel (0.05) ou entiers
                        val_float = float(val)
                    except ValueError:
                        pass # Ignorer valeurs non numériques
                
                # --- TARGET: PAYVAR (Primes 1-5, 13ème) ---
                if info['target'] == 'payvar':
                    # ... PayVar Logic (Currently mostly unused as mapped to PayrollPrime) ...
                    # Only used if PayVar column explicitly imported (13ème Mois?)
                    # But 13th month is now Removed from Template headers...
                    # So likely Unused. Keeping for robustness if old file.
                    if not pv:
                        pv = db.query(models.PayVar).filter(
                            models.PayVar.worker_id == worker.id, 
                            models.PayVar.period == period
                        ).first()
                        if not pv:
                            pv = models.PayVar(worker_id=worker.id, period=period)
                            db.add(pv)
                    
                    final_val = val_float if val_float is not None else 0.0
                    setattr(pv, info['field'], final_val)
                    count_updated += 1
                
                # --- TARGET: PAYROLL PRIME ---
                elif info['target'] == 'payroll_prime':
                    label = info['label']
                    prime_entry = primes_map.get(label)
                    
                    if not prime_entry:
                        # Si vide et n'existe pas, inutile de créer
                        if val_float is None:
                            continue
                            
                        prime_entry = models.PayrollPrime(
                            worker_id=worker.id,
                            period=period,
                            prime_label=label
                        )
                        db.add(prime_entry)
                        primes_map[label] = prime_entry
                    
                    # Update fields
                    if info['type'] == 'nombre':
                        prime_entry.nombre = val_float
                    elif info['type'] == 'base':
                        prime_entry.base = val_float
                    
                    # Clear PayVar legacy simple amount for this prime
                    # (Only if we actually TOUCH IT - implying override intent)
                    if val_float is not None:
                        payvar_field = label_to_field.get(label)
                        if payvar_field:
                             if not pv:
                                 pv = db.query(models.PayVar).filter(
                                     models.PayVar.worker_id == worker.id,
                                     models.PayVar.period == period
                                 ).first()
                             if pv:
                                 setattr(pv, payvar_field, 0.0)
                    
                    # Cleanup: Si tout est NULL, supprimer l'entrée
                    # Includes check for taux (even if not imported)
                    if (prime_entry.nombre is None and 
                        prime_entry.base is None and 
                        prime_entry.taux is None):
                        db.delete(prime_entry)
                        # Remove from map to prevent resurrection
                        if label in primes_map:
                            del primes_map[label]
                        # Don't continue, loop continues to next column
                    
                    count_updated += 1
        
        db.commit()
        
        return {
            "message": "Import terminé avec succès",
            "updated_items": count_updated,
            "errors": errors
        }


    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'import: {str(e)}")


class PrimeValuesOut(BaseModel):
    worker_id: int
    matricule: str
    nom: str
    prenom: str
    prime_13: float
    prime1: float
    prime2: float
    prime3: float
    prime4: float
    prime5: float

class PrimeValuesUpdate(BaseModel):
    prime_13: float
    prime1: float
    prime2: float
    prime3: float
    prime4: float
    prime5: float

@router.get("/values/{payroll_run_id}", response_model=List[PrimeValuesOut])
def get_prime_values(payroll_run_id: int, db: Session = Depends(get_db)):
    # 1. Get PayrollRun
    run = db.query(models.PayrollRun).filter(models.PayrollRun.id == payroll_run_id).first()
    if not run:
        raise HTTPException(404, "Run not found")
        
    period = run.period
    employer_id = run.employer_id
    
    # 2. Get Workers
    workers = db.query(models.Worker).filter(models.Worker.employer_id == employer_id).all()
    
    # 3. Get PayVars
    payvars = db.query(models.PayVar).filter(models.PayVar.period == period).all()
    pv_map = {p.worker_id: p for p in payvars}
    
    result = []
    for w in workers:
        pv = pv_map.get(w.id)
        result.append({
            "worker_id": w.id,
            "matricule": w.matricule,
            "nom": w.nom,
            "prenom": w.prenom,
            "prime_13": pv.prime_13 if pv else 0.0,
            "prime1": pv.prime1 if pv else 0.0,
            "prime2": pv.prime2 if pv else 0.0,
            "prime3": pv.prime3 if pv else 0.0,
            "prime4": pv.prime4 if pv else 0.0,
            "prime5": pv.prime5 if pv else 0.0,
        })
    return result

@router.put("/values/{payroll_run_id}/{worker_id}")
def update_prime_values(
    payroll_run_id: int, 
    worker_id: int, 
    values: PrimeValuesUpdate, 
    db: Session = Depends(get_db)
):
    run = db.query(models.PayrollRun).filter(models.PayrollRun.id == payroll_run_id).first()
    if not run:
        raise HTTPException(404, "Run not found")
        
    period = run.period
    employer_id = run.employer_id
    
    # Get employer to retrieve prime labels
    employer = db.query(models.Employer).filter(models.Employer.id == employer_id).first()
    
    # Update PayVar
    pv = db.query(models.PayVar).filter(models.PayVar.worker_id == worker_id, models.PayVar.period == period).first()
    if not pv:
        pv = models.PayVar(worker_id=worker_id, period=period)
        db.add(pv)
    
    pv.prime_13 = values.prime_13
    pv.prime1 = values.prime1
    pv.prime2 = values.prime2
    pv.prime3 = values.prime3
    pv.prime4 = values.prime4
    pv.prime5 = values.prime5
    
    # LAST MODIFICATION WINS: Delete conflicting PayrollPrime overrides
    # This ensures manual edits take precedence over previous imports
    if employer:
        prime_labels_to_delete = []
        
        # Map prime fields to their labels
        prime_label_map = {
            "prime1": employer.label_prime1 or "Prime 1",
            "prime2": employer.label_prime2 or "Prime 2",
            "prime3": employer.label_prime3 or "Prime 3",
            "prime4": employer.label_prime4 or "Prime 4",
            "prime5": employer.label_prime5 or "Prime 5",
        }
        
        # Collect labels for primes that were just updated
        for field, label in prime_label_map.items():
            prime_labels_to_delete.append(label)
        
        # Also delete "13ème Mois" if it exists as override
        prime_labels_to_delete.append("13ème Mois")
        
        # Delete all PayrollPrime entries for these labels
        db.query(models.PayrollPrime).filter(
            models.PayrollPrime.worker_id == worker_id,
            models.PayrollPrime.period == period,
            models.PayrollPrime.prime_label.in_(prime_labels_to_delete)
        ).delete(synchronize_session=False)
    
    db.commit()
    return {"message": "Updated"}

@router.post("/values/{payroll_run_id}/reset-bulk")
def reset_bulk_prime_values(
    payroll_run_id: int, 
    worker_ids: List[int], 
    db: Session = Depends(get_db)
):
    run = db.query(models.PayrollRun).filter(models.PayrollRun.id == payroll_run_id).first()
    if not run:
        raise HTTPException(404, "Run not found")
        
    period = run.period
    
    # Update PayVars to set primes to 0 for selected workers
    # We must UPDATE.
    
    db.query(models.PayVar).filter(
        models.PayVar.period == period,
        models.PayVar.worker_id.in_(worker_ids)
    ).update({
        "prime_13": 0.0,
        "prime1": 0.0,
        "prime2": 0.0,
        "prime3": 0.0,
        "prime4": 0.0,
        "prime5": 0.0
    }, synchronize_session=False)
    
    db.commit()
    return {"message": "Reset done"}
