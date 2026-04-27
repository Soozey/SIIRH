import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Créer un nouveau classeur
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Import Paie'

# Définir les en-têtes (19 colonnes)
headers = [
    'Matricule',      # 1
    'HSNI_130',       # 2
    'HSI_130',        # 3
    'HSNI_150',       # 4
    'HSI_150',        # 5
    'HMNH',           # 6
    'HMNO',           # 7
    'HMD',            # 8
    'HMJF',           # 9
    'ABSM_J',         # 10
    'ABSM_H',         # 11
    'ABSNR_J',        # 12
    'ABSNR_H',        # 13
    'ABSMP',          # 14
    'ABS1_J',         # 15
    'ABS1_H',         # 16
    'ABS2_J',         # 17
    'ABS2_H',         # 18
    'Avance'          # 19
]

# Ajouter les en-têtes
ws.append(headers)

# Formater les en-têtes
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Ajouter des exemples de données
ws.append(['EMP001', 5.0, 2.0, 3.0, 1.0, 4.0, 2.0, 0.0, 0.0, 0.0, 0.0, 1.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 50000])
ws.append(['EMP002', 0.0, 0.0, 2.0, 0.0, 0.0, 0.0, 0.0, 0.0, 2.0, 0.0, 0.0, 2.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0])

# Ajuster la largeur des colonnes
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Sauvegarder
wb.save('Template_Import_Paie.xlsx')
print(f'✅ Template créé avec {len(headers)} colonnes!')
print(f'Colonnes: {", ".join(headers)}')
