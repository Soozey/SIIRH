import hashlib
import json
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy.orm import Session

from .. import models
from ..config.config import settings
from ..routers.payroll import generate_preview_data
from ..routers.reporting import get_dynamic_journal_columns, get_full_report_data
from ..services.compliance_service import build_contract_checklist, build_contract_snapshot, collect_integrity_issues
from ..services.file_storage import get_upload_root, sanitize_filename_part
from ..services.master_data_service import build_worker_reporting_payload
from ..services.pdf_generation_service import build_report_pdf


DEFAULT_EXPORT_TEMPLATES = [
    {
        "code": "dns_cnaps",
        "type_document": "DNS CNaPS",
        "version": "2026.03",
        "format": "xlsx",
        "mapping": {
            "company_cnaps_id": "employers.cnaps_employer_id",
            "employee_id": "workers.matricule",
            "cnaps_no": "workers.cnaps_num",
            "assiette_cnaps": "payroll.totals.brut",
        },
        "options": {"channels": ["cnaps"], "label": "Declaration nominative des salaires CNaPS"},
    },
    {
        "code": "ostie_smie",
        "type_document": "OSTIE / SMIE",
        "version": "2026.03",
        "format": "xlsx",
        "mapping": {
            "company_ostie_id": "employers.ostie_id",
            "employee_id": "workers.matricule",
            "assiette": "payroll.totals.brut",
            "part_salarie": "payroll.lines.SMIE",
        },
        "options": {"channels": ["ostie"], "label": "Etat OSTIE / SMIE"},
    },
    {
        "code": "irsa_bimestriel",
        "type_document": "IRSA bimestriel",
        "version": "2026.03",
        "format": "xlsx",
        "mapping": {
            "company_nif": "employers.nif",
            "employee_id": "workers.matricule",
            "assiette_irsa": "payroll.totals.net_imposable",
            "irsa_ret": "payroll.totals.irsa",
        },
        "options": {"channels": ["irsa"], "label": "Etat nominatif IRSA"},
    },
    {
        "code": "fmfp",
        "type_document": "FMFP",
        "version": "2026.03",
        "format": "xlsx",
        "mapping": {
            "company_id": "employers.id",
            "base_formation": "payroll.totals.brut",
            "montant_du": "payroll.lines.FMFP",
        },
        "options": {"channels": ["fmfp"], "label": "Versement FMFP"},
    },
    {
        "code": "etat_paie",
        "type_document": "Etat de paie",
        "version": "2026.03",
        "format": "xlsx",
        "mapping": {"source": "reporting.dynamic_journal"},
        "options": {"channels": ["payroll_state"], "label": "Etat de paie"},
    },
    {
        "code": "bilan_social",
        "type_document": "Bilan social annuel",
        "version": "2026.03",
        "format": "xlsx",
        "mapping": {"source": "reporting.social_summary"},
        "options": {"channels": ["social_report"], "label": "Bilan social annuel"},
    },
    {
        "code": "contract_control_bundle",
        "type_document": "Bundle de controle contractuel",
        "version": "2026.03",
        "format": "pdf",
        "mapping": {"source": "contracts.compliance_bundle"},
        "options": {"channels": ["inspection"], "label": "Dossier inspection / conformite"},
    },
]


def _json_dump(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _json_load(value: Any, default: Any):
    if value in (None, ""):
        return default
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return default


def _month_sequence(start_period: str, end_period: str) -> list[str]:
    start = datetime.strptime(start_period, "%Y-%m").date()
    end = datetime.strptime(end_period, "%Y-%m").date()
    cursor = date(start.year, start.month, 1)
    months: list[str] = []
    while cursor <= end:
        months.append(cursor.strftime("%Y-%m"))
        if cursor.month == 12:
            cursor = date(cursor.year + 1, 1, 1)
        else:
            cursor = date(cursor.year, cursor.month + 1, 1)
    return months


def _period_label(start_period: str, end_period: str) -> str:
    return start_period if start_period == end_period else f"{start_period} -> {end_period}"


def _export_directory() -> Path:
    root = get_upload_root(settings.UPLOAD_DIR)
    directory = root / "exports"
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def _save_generated_bytes(*, employer_id: int, template_code: str, extension: str, content: bytes) -> tuple[str, str]:
    directory = _export_directory() / f"employer_{employer_id}" / sanitize_filename_part(template_code)
    directory.mkdir(parents=True, exist_ok=True)
    filename = f"{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{sanitize_filename_part(template_code)}.{extension}"
    path = directory / filename
    path.write_bytes(content)
    checksum = hashlib.sha256(content).hexdigest()
    return str(path), checksum


def _line_amount(preview: dict[str, Any], label_fragment: str, *, patronal: bool = False) -> float:
    total = 0.0
    target = label_fragment.lower()
    for line in preview.get("lines", []):
        label = str(line.get("label", "")).lower()
        if target in label:
            total += float(line.get("montant_pat" if patronal else "montant_sal", 0) or 0)
    return abs(total)


def _payroll_summary_for_worker(db: Session, worker: models.Worker, period: str) -> dict[str, Any]:
    preview = generate_preview_data(worker.id, period, db)
    totals = preview.get("totaux", {})
    brut = float(totals.get("brut", 0) or 0)
    irsa = abs(float(totals.get("irsa", 0) or 0))
    cnaps_sal = _line_amount(preview, "cnaps", patronal=False)
    cnaps_pat = _line_amount(preview, "cnaps", patronal=True)
    smie_sal = _line_amount(preview, "smie", patronal=False)
    smie_pat = _line_amount(preview, "smie", patronal=True)
    fmfp_pat = _line_amount(preview, "fmfp", patronal=True)
    return {
        "preview": preview,
        "brut": brut,
        "irsa": irsa,
        "cnaps_sal": cnaps_sal,
        "cnaps_pat": cnaps_pat,
        "smie_sal": smie_sal,
        "smie_pat": smie_pat,
        "fmfp_pat": fmfp_pat,
        "net": float(totals.get("net", 0) or 0),
        "net_imposable": brut - cnaps_sal - smie_sal,
    }


def _canonical_worker_data(db: Session, worker: models.Worker) -> dict[str, Any]:
    return build_worker_reporting_payload(db, worker)


def ensure_export_templates(db: Session) -> list[models.ExportTemplate]:
    current = {item.code: item for item in db.query(models.ExportTemplate).all()}
    changed = False
    for payload in DEFAULT_EXPORT_TEMPLATES:
        if payload["code"] in current:
            continue
        db.add(
            models.ExportTemplate(
                code=payload["code"],
                type_document=payload["type_document"],
                version=payload["version"],
                format=payload["format"],
                mapping_json=_json_dump(payload["mapping"]),
                options_json=_json_dump(payload["options"]),
                is_active=True,
            )
        )
        changed = True
    if changed:
        db.commit()
    return db.query(models.ExportTemplate).filter(models.ExportTemplate.is_active == True).order_by(models.ExportTemplate.type_document.asc()).all()


def serialize_template(item: models.ExportTemplate) -> dict[str, Any]:
    return {
        "id": item.id,
        "code": item.code,
        "type_document": item.type_document,
        "version": item.version,
        "format": item.format,
        "mapping": _json_load(item.mapping_json, {}),
        "options": _json_load(item.options_json, {}),
        "is_active": item.is_active,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }


def serialize_export_job(item: models.ExportJob) -> dict[str, Any]:
    return {
        "id": item.id,
        "employer_id": item.employer_id,
        "template_id": item.template_id,
        "snapshot_id": item.snapshot_id,
        "requested_by_user_id": item.requested_by_user_id,
        "document_type": item.document_type,
        "start_period": item.start_period,
        "end_period": item.end_period,
        "status": item.status,
        "file_path": item.file_path,
        "checksum": item.checksum,
        "logs": _json_load(item.logs_json, []),
        "errors": _json_load(item.errors_json, []),
        "completed_at": item.completed_at,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }


def serialize_declaration(item: models.StatutoryDeclaration) -> dict[str, Any]:
    return {
        "id": item.id,
        "employer_id": item.employer_id,
        "export_job_id": item.export_job_id,
        "channel": item.channel,
        "period_label": item.period_label,
        "status": item.status,
        "reference_number": item.reference_number,
        "receipt_path": item.receipt_path,
        "totals": _json_load(item.totals_json, {}),
        "metadata": _json_load(item.metadata_json, {}),
        "submitted_at": item.submitted_at,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }


def _build_data_issues(
    db: Session,
    employer: models.Employer,
    workers: list[models.Worker],
    document_code: str,
) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    if document_code == "irsa_bimestriel" and not employer.nif:
        issues.append(
            {
                "severity": "high",
                "issue_type": "missing_company_nif",
                "entity_type": "employer",
                "entity_id": str(employer.id),
                "message": "Le NIF employeur est requis pour l'export IRSA.",
                "details": {},
            }
        )
    for worker in workers:
        canonical = _canonical_worker_data(db, worker)
        worker_name = f"{canonical.get('nom', '')} {canonical.get('prenom', '')}".strip()
        cnaps_number = canonical.get("cnaps_num")
        cin_number = canonical.get("cin")
        if document_code == "dns_cnaps" and not cnaps_number:
            issues.append(
                {
                    "severity": "medium",
                    "issue_type": "missing_cnaps_number",
                    "entity_type": "worker",
                    "entity_id": str(worker.id),
                    "message": "Le numero CNaPS du salarie est manquant pour la DNS.",
                    "details": {"worker_name": worker_name},
                }
            )
        if document_code == "ostie_smie" and not cin_number:
            issues.append(
                {
                    "severity": "medium",
                    "issue_type": "missing_cin",
                    "entity_type": "worker",
                    "entity_id": str(worker.id),
                    "message": "Le CIN du salarie est manquant pour le fichier OSTIE / SMIE.",
                    "details": {"worker_name": worker_name},
                }
            )
    return issues


def _worksheet_title(ws, title: str, subtitle: str = ""):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, size=11)


def _autosize_columns(ws, max_width: int = 36):
    for column_cells in ws.columns:
        first = column_cells[0]
        letter = first.column_letter
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _write_table(ws, start_row: int, headers: list[str], rows: list[dict[str, Any]]):
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row_cursor = start_row + 1
    for row in rows:
        for index, header in enumerate(headers, start=1):
            ws.cell(row=row_cursor, column=index, value=row.get(header))
        row_cursor += 1
    _autosize_columns(ws)


def _preview_dns_cnaps(db: Session, employer: models.Employer, workers: list[models.Worker], start_period: str, end_period: str) -> dict[str, Any]:
    months = _month_sequence(start_period, end_period)
    rows: list[dict[str, Any]] = []
    totals = {"assiette_cnaps": 0.0, "part_salarie": 0.0, "part_employeur": 0.0}
    for worker in workers:
        canonical = _canonical_worker_data(db, worker)
        assiette = 0.0
        part_sal = 0.0
        part_pat = 0.0
        for month in months:
            summary = _payroll_summary_for_worker(db, worker, month)
            assiette += summary["brut"]
            part_sal += summary["cnaps_sal"]
            part_pat += summary["cnaps_pat"]
        totals["assiette_cnaps"] += assiette
        totals["part_salarie"] += part_sal
        totals["part_employeur"] += part_pat
        rows.append(
            {
                "company_cnaps_id": getattr(employer, "cnaps_num", None),
                "period": _period_label(start_period, end_period),
                "employee_id": canonical.get("matricule"),
                "cnaps_no": canonical.get("cnaps_num"),
                "lastname": canonical.get("nom"),
                "firstname": canonical.get("prenom"),
                "hire_date": canonical.get("date_embauche") or None,
                "term_date": canonical.get("date_debauche") or None,
                "assiette_cnaps": round(assiette, 2),
                "part_salarie": round(part_sal, 2),
                "part_employeur": round(part_pat, 2),
                "bank_ov_ref": f"OV-{employer.id}-{start_period.replace('-', '')}",
            }
        )
    return {
        "meta": {
            "document_type": "dns_cnaps",
            "label": "DNS CNaPS",
            "period_label": _period_label(start_period, end_period),
            "company": employer.raison_sociale,
            "totals": totals,
        },
        "columns": [
            "company_cnaps_id",
            "period",
            "employee_id",
            "cnaps_no",
            "lastname",
            "firstname",
            "hire_date",
            "term_date",
            "assiette_cnaps",
            "part_salarie",
            "part_employeur",
            "bank_ov_ref",
        ],
        "rows": rows,
    }


def _preview_ostie_smie(db: Session, employer: models.Employer, workers: list[models.Worker], start_period: str, end_period: str) -> dict[str, Any]:
    months = _month_sequence(start_period, end_period)
    rows: list[dict[str, Any]] = []
    totals = {"assiette": 0.0, "part_salarie": 0.0, "part_employeur": 0.0}
    for worker in workers:
        canonical = _canonical_worker_data(db, worker)
        assiette = 0.0
        part_sal = 0.0
        part_pat = 0.0
        monthly_salaries: list[float] = []
        for month in months:
            summary = _payroll_summary_for_worker(db, worker, month)
            assiette += summary["brut"]
            part_sal += summary["smie_sal"]
            part_pat += summary["smie_pat"]
            monthly_salaries.append(round(summary["brut"], 2))
        totals["assiette"] += assiette
        totals["part_salarie"] += part_sal
        totals["part_employeur"] += part_pat
        rows.append(
            {
                "company_ostie_id": getattr(employer, "smie_num", None),
                "period": _period_label(start_period, end_period),
                "employee_id": canonical.get("matricule"),
                "lastname": canonical.get("nom"),
                "firstname": canonical.get("prenom"),
                "sex": canonical.get("sexe"),
                "birth_date": canonical.get("date_naissance") or None,
                "hire_date": canonical.get("date_embauche") or None,
                "term_date": canonical.get("date_debauche") or None,
                "position": canonical.get("poste"),
                "cnaps_no": canonical.get("cnaps_num"),
                "cin": canonical.get("cin"),
                "salary_month_1": monthly_salaries[0] if len(monthly_salaries) > 0 else 0.0,
                "salary_month_2": monthly_salaries[1] if len(monthly_salaries) > 1 else 0.0,
                "salary_month_3": monthly_salaries[2] if len(monthly_salaries) > 2 else 0.0,
                "total_salary_non_plafonne": round(assiette, 2),
                "total_salary_plafonne": round(assiette, 2),
                "part_employeur": round(part_pat, 2),
                "part_travailleur": round(part_sal, 2),
            }
        )
    return {
        "meta": {
            "document_type": "ostie_smie",
            "label": "OSTIE / SMIE",
            "period_label": _period_label(start_period, end_period),
            "company": employer.raison_sociale,
            "totals": totals,
        },
        "columns": [
            "company_ostie_id",
            "period",
            "employee_id",
            "lastname",
            "firstname",
            "sex",
            "birth_date",
            "hire_date",
            "term_date",
            "position",
            "cnaps_no",
            "cin",
            "salary_month_1",
            "salary_month_2",
            "salary_month_3",
            "total_salary_non_plafonne",
            "total_salary_plafonne",
            "part_employeur",
            "part_travailleur",
        ],
        "rows": rows,
    }


def _preview_irsa(db: Session, employer: models.Employer, workers: list[models.Worker], start_period: str, end_period: str) -> dict[str, Any]:
    months = _month_sequence(start_period, end_period)
    rows: list[dict[str, Any]] = []
    totals = {"assiette_irsa": 0.0, "irsa_ret": 0.0}
    for worker in workers:
        canonical = _canonical_worker_data(db, worker)
        assiette = 0.0
        irsa = 0.0
        for month in months:
            summary = _payroll_summary_for_worker(db, worker, month)
            assiette += summary["net_imposable"]
            irsa += summary["irsa"]
        totals["assiette_irsa"] += assiette
        totals["irsa_ret"] += irsa
        rows.append(
            {
                "company_nif": employer.nif,
                "period": _period_label(start_period, end_period),
                "employee_id": canonical.get("matricule"),
                "lastname": canonical.get("nom"),
                "firstname": canonical.get("prenom"),
                "assiette_irsa": round(assiette, 2),
                "irsa_ret": round(irsa, 2),
                "payment_mode": canonical.get("mode_paiement"),
            }
        )
    return {
        "meta": {
            "document_type": "irsa_bimestriel",
            "label": "IRSA bimestriel",
            "period_label": _period_label(start_period, end_period),
            "company": employer.raison_sociale,
            "totals": totals,
        },
        "columns": ["company_nif", "period", "employee_id", "lastname", "firstname", "assiette_irsa", "irsa_ret", "payment_mode"],
        "rows": rows,
    }


def _preview_fmfp(db: Session, employer: models.Employer, workers: list[models.Worker], start_period: str, end_period: str) -> dict[str, Any]:
    months = _month_sequence(start_period, end_period)
    rows: list[dict[str, Any]] = []
    total_base = 0.0
    total_amount = 0.0
    rate = 0.01
    for month in months:
        month_base = 0.0
        month_amount = 0.0
        for worker in workers:
            summary = _payroll_summary_for_worker(db, worker, month)
            month_base += summary["brut"]
            month_amount += summary["fmfp_pat"]
        total_base += month_base
        total_amount += month_amount
        rows.append(
            {
                "company_id": employer.id,
                "period": month,
                "base_formation": round(month_base, 2),
                "taux_employeur": rate,
                "montant_du": round(month_amount, 2),
                "justif_ref": f"FMFP-{employer.id}-{month.replace('-', '')}",
            }
        )
    return {
        "meta": {
            "document_type": "fmfp",
            "label": "FMFP",
            "period_label": _period_label(start_period, end_period),
            "company": employer.raison_sociale,
            "totals": {"base_formation": total_base, "montant_du": total_amount},
        },
        "columns": ["company_id", "period", "base_formation", "taux_employeur", "montant_du", "justif_ref"],
        "rows": rows,
    }


def _preview_payroll_state(db: Session, employer: models.Employer, start_period: str, end_period: str) -> dict[str, Any]:
    columns = get_dynamic_journal_columns(employer.id, db)
    rows = get_full_report_data(employer.id, start_period, end_period, columns, db, filters={}, viewer=None)
    visible_rows = [{k: v for k, v in row.items() if not str(k).startswith("_")} for row in rows]
    return {
        "meta": {
            "document_type": "etat_paie",
            "label": "Etat de paie",
            "period_label": _period_label(start_period, end_period),
            "company": employer.raison_sociale,
            "row_count": len(visible_rows),
        },
        "columns": list(visible_rows[0].keys()) if visible_rows else columns,
        "rows": visible_rows,
    }


def _preview_social_report(db: Session, employer: models.Employer, workers: list[models.Worker], start_period: str, end_period: str) -> dict[str, Any]:
    months = _month_sequence(start_period, end_period)
    payroll_mass = 0.0
    net_total = 0.0
    irsa_total = 0.0
    overtime_estimate = 0.0
    for worker in workers:
        for month in months:
            summary = _payroll_summary_for_worker(db, worker, month)
            payroll_mass += summary["brut"]
            net_total += summary["net"]
            irsa_total += summary["irsa"]
            preview = summary["preview"]
            overtime_estimate += _line_amount(preview, "130", patronal=False) + _line_amount(preview, "150", patronal=False)

    incidents = db.query(models.SstIncident).filter(models.SstIncident.employer_id == employer.id).count()
    active_workers = [worker for worker in workers if not worker.date_debauche]
    rows = [
        {"metric": "effectif_total", "value": len(workers), "description": "Salaries suivis sur la periode"},
        {"metric": "effectif_actif", "value": len(active_workers), "description": "Salaries actifs"},
        {"metric": "masse_salariale_brute", "value": round(payroll_mass, 2), "description": "Somme des bruts lus en paie"},
        {"metric": "net_total", "value": round(net_total, 2), "description": "Somme des nets a payer"},
        {"metric": "irsa_total", "value": round(irsa_total, 2), "description": "IRSA cumule sur la periode"},
        {"metric": "overtime_estimate", "value": round(overtime_estimate, 2), "description": "Approximation monetaire des HS majorations"},
        {"metric": "sst_incidents", "value": incidents, "description": "Incidents SST enregistres"},
    ]
    return {
        "meta": {
            "document_type": "bilan_social",
            "label": "Bilan social annuel",
            "period_label": _period_label(start_period, end_period),
            "company": employer.raison_sociale,
        },
        "columns": ["metric", "value", "description"],
        "rows": rows,
    }


def _preview_contract_bundle(db: Session, employer: models.Employer, workers: list[models.Worker], start_period: str, end_period: str) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    for worker in workers:
        canonical = _canonical_worker_data(db, worker)
        contract = (
            db.query(models.CustomContract)
            .filter(models.CustomContract.worker_id == worker.id)
            .order_by(models.CustomContract.updated_at.desc())
            .first()
        )
        if not contract:
            continue
        checklist = build_contract_checklist(contract, worker)
        rows.append(
            {
                "worker_id": worker.id,
                "matricule": canonical.get("matricule"),
                "worker_name": f"{canonical.get('nom', '')} {canonical.get('prenom', '')}".strip(),
                "contract_id": contract.id,
                "contract_title": contract.title,
                "status": "ok" if all(item["status"] == "ok" for item in checklist) else "a_corriger",
                "missing_items": ", ".join(item["label"] for item in checklist if item["status"] != "ok"),
                "snapshot": build_contract_snapshot(contract, worker),
            }
        )
    return {
        "meta": {
            "document_type": "contract_control_bundle",
            "label": "Bundle controle contractuel",
            "period_label": _period_label(start_period, end_period),
            "company": employer.raison_sociale,
        },
        "columns": ["matricule", "worker_name", "contract_title", "status", "missing_items"],
        "rows": rows,
    }


def build_export_preview(
    db: Session,
    *,
    employer_id: int,
    template_code: str,
    start_period: str,
    end_period: str,
) -> dict[str, Any]:
    templates = {item.code: item for item in ensure_export_templates(db)}
    template = templates.get(template_code)
    if not template:
        raise ValueError("Unknown export template")

    employer = db.query(models.Employer).filter(models.Employer.id == employer_id).first()
    if not employer:
        raise ValueError("Employer not found")
    workers = db.query(models.Worker).filter(models.Worker.employer_id == employer_id).all()
    worker_sort_index = {item.id: _canonical_worker_data(db, item) for item in workers}
    workers.sort(
        key=lambda item: (
            str(worker_sort_index[item.id].get("nom") or ""),
            str(worker_sort_index[item.id].get("prenom") or ""),
            str(worker_sort_index[item.id].get("matricule") or ""),
        )
    )

    if template_code == "dns_cnaps":
        payload = _preview_dns_cnaps(db, employer, workers, start_period, end_period)
    elif template_code == "ostie_smie":
        payload = _preview_ostie_smie(db, employer, workers, start_period, end_period)
    elif template_code == "irsa_bimestriel":
        payload = _preview_irsa(db, employer, workers, start_period, end_period)
    elif template_code == "fmfp":
        payload = _preview_fmfp(db, employer, workers, start_period, end_period)
    elif template_code == "etat_paie":
        payload = _preview_payroll_state(db, employer, start_period, end_period)
    elif template_code == "bilan_social":
        payload = _preview_social_report(db, employer, workers, start_period, end_period)
    elif template_code == "contract_control_bundle":
        payload = _preview_contract_bundle(db, employer, workers, start_period, end_period)
    else:
        raise ValueError("Template not implemented")

    issues = _build_data_issues(db, employer, workers, template_code)
    issues.extend(collect_integrity_issues(db, employer_id))
    return {
        "template_code": template.code,
        "document_type": template.type_document,
        "format": template.format,
        "meta": payload["meta"],
        "columns": payload["columns"],
        "rows": payload["rows"],
        "issues": issues,
    }


def _render_xlsx_from_preview(preview: dict[str, Any]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Export"
    _worksheet_title(worksheet, preview["meta"]["label"], f"PÃ©riode: {preview['meta']['period_label']}")
    _write_table(worksheet, 4, preview["columns"], preview["rows"])

    if preview["template_code"] == "irsa_bimestriel":
        worksheet.title = "EtatNominatif"
        bordereau = workbook.create_sheet("Bordereau")
        _worksheet_title(
            worksheet,
            "ETAT NOMINATIF DES TRAITEMENTS SALAIRES ET ASSIMILES PAYES",
            f"Entreprise: {preview['meta']['company']}",
        )
        _write_table(worksheet, 6, preview["columns"], preview["rows"])
        _worksheet_title(bordereau, "BORDEREAU DE VERSEMENT IRSA", f"PÃ©riode: {preview['meta']['period_label']}")
        totals = preview["meta"].get("totals", {})
        bordereau["A4"] = "Assiette IRSA"
        bordereau["B4"] = round(float(totals.get("assiette_irsa", 0) or 0), 2)
        bordereau["A5"] = "IRSA retenue"
        bordereau["B5"] = round(float(totals.get("irsa_ret", 0) or 0), 2)
        _autosize_columns(bordereau)

    if preview["template_code"] == "bilan_social":
        synthese = workbook.create_sheet("Synthese")
        _worksheet_title(synthese, "Bilan social annuel", preview["meta"]["company"])
        _write_table(synthese, 4, preview["columns"], preview["rows"])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _render_pdf_from_preview(preview: dict[str, Any]) -> bytes:
    rows = []
    for row in preview["rows"]:
        rows.append({key: row.get(key) for key in preview["columns"]})
    return build_report_pdf(
        title=preview["meta"]["label"],
        subtitle=f"{preview['meta']['company']} | {preview['meta']['period_label']}",
        columns=preview["columns"],
        rows=rows,
    )


def generate_export_job(
    db: Session,
    *,
    employer_id: int,
    template_code: str,
    start_period: str,
    end_period: str,
    requested_by: Optional[models.AppUser],
) -> models.ExportJob:
    templates = {item.code: item for item in ensure_export_templates(db)}
    template = templates.get(template_code)
    if not template:
        raise ValueError("Unknown export template")

    preview = build_export_preview(
        db,
        employer_id=employer_id,
        template_code=template_code,
        start_period=start_period,
        end_period=end_period,
    )
    snapshot = models.ReportingSnapshot(
        employer_id=employer_id,
        snapshot_type=template_code,
        start_period=start_period,
        end_period=end_period,
        source_hash=hashlib.sha256(_json_dump(preview["rows"]).encode("utf-8")).hexdigest(),
        data_json=_json_dump(preview),
        created_by_user_id=requested_by.id if requested_by else None,
    )
    db.add(snapshot)
    db.flush()

    job = models.ExportJob(
        employer_id=employer_id,
        template_id=template.id,
        snapshot_id=snapshot.id,
        requested_by_user_id=requested_by.id if requested_by else None,
        document_type=template.type_document,
        start_period=start_period,
        end_period=end_period,
        status="generated",
        logs_json=_json_dump([f"Snapshot {snapshot.id} created", f"Template {template.code} used"]),
        errors_json=_json_dump([]),
        completed_at=datetime.now(timezone.utc),
    )
    db.add(job)
    db.flush()

    content = _render_pdf_from_preview(preview) if template.format == "pdf" else _render_xlsx_from_preview(preview)
    extension = "pdf" if template.format == "pdf" else "xlsx"
    file_path, checksum = _save_generated_bytes(
        employer_id=employer_id,
        template_code=template.code,
        extension=extension,
        content=content,
    )
    job.file_path = file_path
    job.checksum = checksum

    channels = _json_load(template.options_json, {}).get("channels", [])
    for channel in channels:
        db.add(
            models.StatutoryDeclaration(
                employer_id=employer_id,
                export_job_id=job.id,
                channel=channel,
                period_label=_period_label(start_period, end_period),
                status="generated",
                totals_json=_json_dump(preview["meta"].get("totals", {})),
                metadata_json=_json_dump({"template_code": template.code, "document_type": template.type_document}),
            )
        )

    db.commit()
    db.refresh(job)
    return job


def download_path_for_job(job: models.ExportJob) -> Path:
    if not job.file_path:
        raise FileNotFoundError("No generated file")
    path = Path(job.file_path)
    if not path.exists():
        raise FileNotFoundError(str(path))
    return path


